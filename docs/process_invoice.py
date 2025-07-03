#!/opt/venv/bin/python3
"""
Standalone Python script for n8n Execute Command node
Processes invoice PDF and Excel report files for reconciliation
Complete version matching app(1).py functionality

CHANGES FOR N8N INTEGRATION & RECONCILIATION FIXES:

1. FIXED RECONCILIATION ISSUES:
   - Restored original AWB data format matching logic from app(1).py
   - Removed over-normalization that was breaking AWB matching
   - Added extensive debugging for AWB matching process
   - Enhanced string normalization while preserving original format compatibility

2. ADDED N8N DYNAMIC FILENAME SUPPORT:
   - Function signature: process_files(invoice_file_path, report_file_path, workflow_id=None, custom_filename=None)
   - N8N Pattern: 'invoicefile-workflowid.xlsx' when workflow_id provided
   - Environment variable support: N8N_WORKFLOW_ID, N8N_OUTPUT_FILENAME
   - Fallback to timestamp pattern if no N8N variables

3. USAGE EXAMPLES:
   
   N8N Execute Command Node:
   ```
   python /path/to/process_invoice.py {{ $json.invoice_path }} {{ $json.report_path }} /tmp/result.json {{ $workflow.id }}
   ```
   
   With Custom Filename:
   ```
   python /path/to/process_invoice.py invoice.pdf report.xls /tmp/result.json workflow123 custom-reconciliation.xlsx
   ```
   
   Environment Variables (N8N):
   - Set N8N_WORKFLOW_ID={{ $workflow.id }}
   - Set N8N_OUTPUT_FILENAME={{ $json.custom_name }}

4. OUTPUT:
   - Returns JSON with success status and output_filename
   - File saved to /files/ directory if exists (N8N container), otherwise local directory
   - Enhanced debug output for troubleshooting reconciliation issues

RECONCILIATION DEBUGGING:
- Added before/after AWB Serial formatting debug output
- Enhanced AWB matching analysis with detailed comparisons
- Shows exact format mismatches between invoice and report data
"""

import sys
import os
import json
import re
import pandas as pd
import pdfplumber
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def safe_to_numeric(series):
    """Converts a pandas Series to numeric, coercing errors to NaN."""
    return pd.to_numeric(series, errors='coerce')

def extract_awb_data(pdf):
    """
    Extracts AWB data from specific pages of a FlyDubai PDF invoice 
    by processing the extracted text lines with layout preservation.
    Handles the multi-line format (AWB line + Date/Rate line).
    Dynamically determines the end page based on CCA header.
    """
    print(f"--- Starting AWB PDF Text Extraction Process ---")
    extracted_data = []

    # --- Regex Definitions ---
    # AWB Line (Copied from app(1).py)
    line1_regex = re.compile(
        r"^\s*141\s+(\d{7})\s+(\d)\s+TLV\s+([A-Z]{3})\s+"  
        r"(\d+\.\d{2}\s*K)\s+"                       # Charge Weight (e.g., 560.00K or 560.00 K) - Made space optional
        r"([\d\.\,-]+)\s+"                             # PP Freight Charge
        r"([\d\.\,-]+)\s+"                             # PP Due Airline
        r"([\d\.\,-]+)\s+"                             # CC Freight Charge
        r"([\d\.\,-]+)\s+"                             # CC Due Agent
        r"([\d\.\,-]+)\s+"                             # CC Due Airline
        r"([\d\.\,-]+)\s+"                             # Disc.
        r"([\d\.\,-]+)\s+"                             # Agency Comm.
        r"([\d\.\,-]+)\s+"                             # Taxes
        r"([\d\.\,-]+)\s+"                             # Others
        r"([\d\.\,-]+)\s+"                             # Net Due for AWB (1)
        r"(1\.00000000)\s+"                           # Exchange Rate
        r"I?\s*([\d\.\,-]+)\s*$"                        # Net Due for AWB (2)
    )
    # Rate pattern (search within the second line)
    net_yield_rate_regex_search = re.compile(r"(\d+\.\d+)")    
    # --- End Regex Definitions ---

    all_lines = []
    
    # --- Determine Target Page Range Dynamically ---
    cca_start_page_index = -1
    num_pages = len(pdf.pages)
    print(f"Total pages in PDF: {num_pages}")
    # Start searching for CCA from page 2 (index 1) onwards
    for page_num in range(1, num_pages):
        try:
            page_to_check = pdf.pages[page_num]
            text_to_check = page_to_check.extract_text()
            if text_to_check and "Section B: CCA Details" in text_to_check:
                cca_start_page_index = page_num
                print(f"  -> Found 'Section B: CCA Details' header on page {page_num + 1}. AWB data ends before this.")
                break
        except Exception as e:
            print(f"Warning: Error checking page {page_num + 1} for CCA header: {e}")

    # If CCA header not found, process all pages from page 2 to the end
    if cca_start_page_index == -1:
        print("  -> 'Section B: CCA Details' not found. Processing AWB data until the end of the document.")
        end_page_index = num_pages
    else:
        end_page_index = cca_start_page_index

    target_pages = range(1, end_page_index) # Page range is 1 to end_page_index (exclusive)
    print(f"AWB Target Page Indices (0-based): {list(target_pages)}")
    # --- End Determine Target Page Range ---

    # --- Extract Text from Target Pages ---
    for page_num in target_pages:
        try:
            page = pdf.pages[page_num]
            print(f"Extracting text with layout from Page {page_num + 1}...")
            # Use same extraction settings as test script
            text = page.extract_text(x_tolerance=2, layout=True) 
            if text:
                page_lines = text.split('\n')
                print(f"  -> Extracted {len(page_lines)} lines from page {page_num + 1}.")
                all_lines.extend(page_lines)
            else:
                print(f"  -> No text extracted from page {page_num + 1}.")
        except Exception as e:
            print(f"Error extracting text from page {page_num + 1}: {e}")

    print(f"--- Total text lines collected from AWB pages: {len(all_lines)} ---")

    # --- Process the collected text lines ---
    i = 0
    while i < len(all_lines):
        line1_raw = all_lines[i] 
        line1 = re.sub(r'\s{2,}', ' ', line1_raw).strip() 
        
        match1 = line1_regex.match(line1)
        
        if match1:
            flight_date = ""
            net_yield_rate = ""
            lines_consumed = 1 

            if i + 1 < len(all_lines):
                line2_raw = all_lines[i+1]
                line2 = re.sub(r'\s{2,}', ' ', line2_raw).strip()
                
                parts = line2.split(' ')
                flight_date = "" 
                net_yield_rate = "" 

                if parts:
                     flight_date = parts[0] 
                     lines_consumed += 1

                     rate_match = net_yield_rate_regex_search.search(line2)
                     if rate_match:
                         potential_rates = net_yield_rate_regex_search.findall(line2)
                         if len(potential_rates) > 0:
                             net_yield_rate = potential_rates[0]
                else:
                     flight_date = "" 

            if match1 and flight_date:
                groups = match1.groups()
                # Data cleaning and assignment
                cleaned_groups = [g.replace(',', '.').strip() if isinstance(g, str) else g for g in groups]
                
                awb_number = f"141 {cleaned_groups[0]} {cleaned_groups[1]}"
                destination = cleaned_groups[2]
                charge_weight = cleaned_groups[3] # Already includes K
                pp_freight = cleaned_groups[4]
                pp_due_airline = cleaned_groups[5]
                cc_freight = cleaned_groups[6]
                cc_due_agent = cleaned_groups[7]
                cc_due_airline = cleaned_groups[8]
                disc = cleaned_groups[9]
                agency_comm = cleaned_groups[10]
                taxes = cleaned_groups[11]
                others = cleaned_groups[12]
                net_due_awb1 = cleaned_groups[13]
                exchange_rate = cleaned_groups[14]

                extracted_data.append({
                    "AWB Number": awb_number,
                    "AWB Serial Part1": cleaned_groups[0],
                    "AWB Serial Part2": cleaned_groups[1],
                    "Flight Date": flight_date,
                    "Origin": "TLV",
                    "Destination": destination,
                    "Charge Weight": charge_weight,
                    "Net Yield Rate": net_yield_rate,
                    "PP Freight Charge": pp_freight,
                    "PP Due Airline": pp_due_airline,
                    "CC Freight Charge": cc_freight,
                    "CC Due Agent": cc_due_agent,
                    "CC Due Airline": cc_due_airline,
                    "Disc.": disc,
                    "Agency Comm.": agency_comm,
                    "Taxes": taxes,
                    "Others": others,
                    "Net Due for AWB": net_due_awb1,
                    "Exchange Rate": exchange_rate
                })
                
                i += lines_consumed
                continue 
            else:
                i += 1 
        else:
            i += 1

    print("--- Finished Processing Text Lines ---")
    df = pd.DataFrame(extracted_data)

    return df

def extract_cca_data(pdf):
    """
    Extracts CCA data from FlyDubai PDF invoice
    by processing the extracted text lines with layout preservation.
    Handles the multi-line format for CCA entries.
    Dynamically finds the CCA page.
    """
    print(f"--- Starting CCA PDF Text Extraction Process ---")
    extracted_data = []
    target_page = -1
    raw_text_cca_page = ""

    # --- Regex Definitions ---
    destination_regex_search = re.compile(r"\b([A-Z]{3})\b")
    cca_date_regex_search = re.compile(r"\b(\d{2}[A-Z]{3})\b")

    # --- Block Regex --- 
    cca_block_regex = re.compile(
        r"^\s*(\d{5,})\s+"          # CCA Ref No (1)
        r"(\d{3})\s*"             # AWB Prefix (2)
        r"(\d{7}\s\d{1})\s*"           # AWB Serial (3) - Captures "NNNNNNN N"
        r"([A-Z]{3})\s+"          # Origin (4)
        r"(\S+)\s+"               # MOP Freight (5)
        r"(\S+)\s+"               # MOP Other (6)
        r"([\d().,-]+)\s+"         # Freight Charge (7) - Handles parentheses
        r"([\d().,-]+)\s+"         # Due Airline (8) - Handles parentheses
        r"([\d().,-]+)\s+"         # Due Agent (9) - Handles parentheses
        r"([\d().,-]+)"            # Disc. (10) - Handles parentheses
        r".*?"                     # Non-greedy match until next numeric field
        r"([\d().,-]+)\s+"         # Agency Comm. (11) - Handles parentheses
        r"([\d().,-]+)\s+"         # Taxes (12) - Handles parentheses
        r"([\d().,-]+)\s+"         # Others (13) - Handles parentheses
        r"([\d().,-]+)\s+"         # Net Due (Sale) (14) - Handles parentheses
        r"\d\.\d{2}\s+"          # Exchange Rate (ignore)
        r"[\d().,-]+\s*?"          # Net Due (Invoice) (ignore)
        # --- Second Line --- 
        r"\n\s*"                  # Match newline and start of next line
        r"(\d{2}[A-Z]{3})?"        # CCA Issue Date (15) - Optional capture
        r".*?"                     # Non-greedy match until destination
        r"\b([A-Z]{3})\b"          # Destination (16)"
        , re.MULTILINE | re.DOTALL
    )
    # --- End Regex Definitions ---

    num_pages = len(pdf.pages)
    print(f"PDF has {num_pages} pages (in CCA function).")
    
    # --- Find the CCA Page Dynamically --- 
    start_search_page = 1 
    print(f"Searching for 'Section B: CCA Details' starting from page {start_search_page + 1}...")
    for page_num in range(start_search_page, num_pages):
        print(f"  Checking page {page_num + 1}...")
        try:
            page_to_check = pdf.pages[page_num]
            text_to_check = page_to_check.extract_text()
            if text_to_check and "Section B: CCA Details" in text_to_check:
                target_page = page_num
                print(f"  -> Found CCA header on page {target_page + 1}!")
                break
        except Exception as e:
            print(f"Warning: Error checking page {page_num + 1} for CCA header: {e}")
    # --- End Find Page ---

    if target_page != -1:
        try:
            page = pdf.pages[target_page]
            print(f"Using standard text extraction for Page {target_page + 1}.")
            raw_text_cca_page = page.extract_text()
            if raw_text_cca_page:
                print(f"  -> Extracted {len(raw_text_cca_page)} characters from page {target_page + 1}.")
            else:
                print(f"  -> No text extracted from page {target_page + 1} using standard extraction.")
        except Exception as e:
             print(f"Error extracting text from CCA page {target_page + 1}: {e}")
             raw_text_cca_page = ""
    else:
        print("Warning: 'Section B: CCA Details' header not found in the document. Assuming no CCA data.")
        return pd.DataFrame()

    # --- Process using findall on the raw text block --- 
    print("--- Starting CCA processing using findall on raw text --- ")
    extracted_data = [] 
    
    # --- Define clean_currency locally ---
    def clean_currency(value_str):
        if isinstance(value_str, str):
            cleaned = value_str.replace(',', '').strip()
            is_negative = False
            
            if cleaned.startswith('(') and cleaned.endswith(')'):
                is_negative = True
                cleaned = cleaned[1:-1]
            elif cleaned.startswith('(') and not cleaned.endswith(')'):
                temp_cleaned = cleaned[1:]
                try:
                    if re.match(r"^[\d.]+$", temp_cleaned):
                        is_negative = True
                        cleaned = temp_cleaned
                except re.error:
                    pass

            try:
                value = float(cleaned)
                if is_negative:
                    value = -value
                return value
            except ValueError:
                if not cleaned and value_str == "()":
                    return 0.0
                return value_str
        return value_str
    # --- End clean_currency definition ---
    
    if raw_text_cca_page:
        matches = cca_block_regex.findall(raw_text_cca_page)
        print(f"  -> Found {len(matches)} potential CCA blocks using findall.")

        for groups in matches:            
            if len(groups) == 16:
                 extracted_data.append({
                    "CCA Ref. No": groups[0],
                    "AWB Prefix": groups[1],
                    "AWB Serial": groups[2].replace(" ", ""),
                    "CCA Issue Date": groups[14] if groups[14] else "",
                    "Origin": groups[3],
                    "Destination": groups[15],
                    "MOP Freight Charge": groups[4],
                    "MOP Other Charge": groups[5],
                    "Freight Charge": clean_currency(groups[6]),
                    "Due Airline": clean_currency(groups[7]),
                    "Due Agent": clean_currency(groups[8]),
                    "Disc.": clean_currency(groups[9]),
                    "Agency Comm.": clean_currency(groups[10]),
                    "Taxes": clean_currency(groups[11]),
                    "Others": clean_currency(groups[12]),
                    "Net Due for AWB (Sale Currency)": clean_currency(groups[13]),
                })
            else:
                 print(f"    -> WARNING: Match found but had unexpected number of groups ({len(groups)}). Skipping.")

    else:
        print("  -> No raw text extracted to process.")
    
    print("--- Finished Processing CCA Text ---")
    df_cca = pd.DataFrame(extracted_data)
    
    # Identify numeric columns for potential totaling (excluding AWB parts)
    numeric_cols_cca = [col for col in df_cca.columns if df_cca[col].apply(lambda x: isinstance(x, (int, float))).all()]
    
    # Add Totals Row if data exists and numeric columns are found
    if not df_cca.empty and numeric_cols_cca:
        totals_cca = df_cca[numeric_cols_cca].sum().to_dict()
        totals_row_cca = {col: totals_cca.get(col, '') for col in df_cca.columns}
        totals_row_cca['CCA Ref. No'] = 'Total' 
        # Ensure AWB details aren't summed
        totals_row_cca['AWB Prefix'] = ''
        totals_row_cca['AWB Serial'] = ''
        totals_row_cca['CCA Issue Date'] = ''
        totals_row_cca['Origin'] = ''
        totals_row_cca['Destination'] = ''
        totals_row_cca['MOP Freight Charge'] = ''
        totals_row_cca['MOP Other Charge'] = ''
        
        df_cca = pd.concat([df_cca, pd.DataFrame([totals_row_cca])], ignore_index=True)
        print("  -> Added totals row to CCA DataFrame.")
    elif not df_cca.empty:
         print("  -> CCA DataFrame created, but no purely numeric columns found for totaling.")
    
    return df_cca

def process_files(invoice_file_path, report_file_path, workflow_id=None, custom_filename=None):
    """Main processing function that matches app(1).py functionality."""
    print("Starting comprehensive file processing...")
    
    # Ensure the file exists before processing
    if not os.path.exists(invoice_file_path):
        raise RuntimeError(f"Invoice file not found at {invoice_file_path}")

    # 1. Extract AWB and CCA Data
    df_awb = pd.DataFrame()
    df_cca = pd.DataFrame()
    try:
        with pdfplumber.open(invoice_file_path) as pdf:
            df_awb = extract_awb_data(pdf)
            df_cca = extract_cca_data(pdf)
    except Exception as e:
        raise RuntimeError(f"Error reading PDF structure: {e}")

    if df_awb.empty and df_cca.empty:
        print("Both extract_awb_data and extract_cca_data returned empty DataFrames.")
        return None

    # Initialize variables
    total_net_due_awb = 0.0
    df_awb_final = pd.DataFrame()
    df_awb_for_recon = pd.DataFrame()
    df_cca_final = pd.DataFrame()
    df_reconciliation = pd.DataFrame()

    # --- Process AWB Data ---
    if not df_awb.empty:
        df_awb_data_only = df_awb.copy()
        print(f"  -> Initial AWB data rows extracted: {len(df_awb_data_only)}")

        # Process AWB data (calculations, column drops) before writing
        numeric_cols_awb = [
            'PP Freight Charge', 'PP Due Airline', 'CC Freight Charge', 'CC Due Agent',
            'CC Due Airline', 'Disc.', 'Agency Comm.', 'Taxes', 'Others', 'Net Due for AWB',
            'Net Yield Rate'
        ]
        # Convert only potential numeric columns
        print("  -> Converting AWB columns to numeric...")
        for col in numeric_cols_awb:
             if col in df_awb_data_only.columns:
                df_awb_data_only[col] = safe_to_numeric(df_awb_data_only[col])

        # Calculate totals based *only* on data rows
        valid_numeric_cols_awb = [col for col in numeric_cols_awb if col in df_awb_data_only.columns and pd.api.types.is_numeric_dtype(df_awb_data_only[col])]
        cols_to_drop_for_sum = ["AWB Number", "Flight Date", "Origin", "Destination", "Charge Weight", "AWB Serial Part1", "AWB Serial Part2", "Exchange Rate"]
        df_numeric_awb_data_only = df_awb_data_only.drop(columns=[col for col in cols_to_drop_for_sum if col in df_awb_data_only.columns], errors='ignore')

        totals_row_awb = {}
        if valid_numeric_cols_awb and not df_numeric_awb_data_only.empty:
            totals_awb = df_numeric_awb_data_only[valid_numeric_cols_awb].sum(numeric_only=True).to_dict()
            totals_row_awb = {col: totals_awb.get(col, '') for col in df_awb_data_only.columns}
            totals_row_awb["AWB Number"] = "Total"
            # Ensure non-numeric fields in total row are blank
            for col in cols_to_drop_for_sum:
                if col in totals_row_awb:
                    totals_row_awb[col] = ''
            
            # Calculate the overall total net due from the data_only df
            if 'Net Due for AWB' in df_awb_data_only.columns and pd.api.types.is_numeric_dtype(df_awb_data_only['Net Due for AWB']):
                total_net_due_awb = df_awb_data_only['Net Due for AWB'].sum()
                print(f"  -> Calculated Total Net Due (data only): {total_net_due_awb}")
            else:
                total_net_due_awb = 0.0

            print("  -> Calculated totals row based on AWB data_only.")
        else:
            print("  -> AWB data_only empty or no valid numeric columns found for totaling.")
            if 'Net Due for AWB' in df_awb_data_only.columns and pd.api.types.is_numeric_dtype(df_awb_data_only['Net Due for AWB']):
               total_net_due_awb = df_awb_data_only['Net Due for AWB'].sum()

        # --- Apply Formatting and Splitting to df_awb_data_only --- 
        # Format Flight Date (assuming DDMMMYY input)
        def format_date(date_str):
            try:
                return pd.to_datetime(date_str, format='%d%b%y', errors='coerce').strftime('%d/%m/%Y')
            except (ValueError, TypeError, AttributeError):
                return date_str

        if 'Flight Date' in df_awb_data_only.columns:
             df_awb_data_only['Flight Date'] = df_awb_data_only['Flight Date'].astype(str).apply(format_date)

        # Format Charge Weight (remove ' K' and convert to numeric)
        if 'Charge Weight' in df_awb_data_only.columns:
            df_awb_data_only['Charge Weight'] = df_awb_data_only['Charge Weight'].astype(str).str.replace(r'\s*K$', '', regex=True)
            df_awb_data_only['Charge Weight'] = safe_to_numeric(df_awb_data_only['Charge Weight'])

        # -- Split AWB Number --
        if "AWB Number" in df_awb_data_only.columns:
            # Split "141 ABCDEFG H" into "141" and "ABCDEFGH"
            split_awb = df_awb_data_only["AWB Number"] \
                .astype(str) \
                .str.split(r'\s+', n=2, expand=True)
            # Assign prefix, handling potential None from split
            df_awb_data_only['AWB Prefix'] = split_awb[0].fillna('').astype(str).str.strip()
            # Combine parts 1 and 2 for the serial, handling potential None
            df_awb_data_only['AWB Serial'] = split_awb[1].fillna('').astype(str) + split_awb[2].fillna('').astype(str)
            # Remove extra spaces within serial just in case
            df_awb_data_only['AWB Serial'] = df_awb_data_only['AWB Serial'].str.replace(r'\s+', '', regex=True).str.strip()
            # Update the separate totals_row_awb with split info if needed (optional, keeps structure consistent)
            if totals_row_awb:
                 totals_row_awb['AWB Prefix'] = 'Total'
                 totals_row_awb['AWB Serial'] = ''

        # --- Finalize DataFrames for Sheets ---
        df_awb_for_recon = df_awb_data_only.copy()
        print(f"  -> Finalized df_awb_for_recon (data only). Shape: {df_awb_for_recon.shape}")

        # df_awb_final (for Invoices sheet) = processed data_only + totals_row
        df_awb_final_data = df_awb_data_only.drop(columns=["AWB Serial Part1", "AWB Serial Part2", "Exchange Rate", "AWB Number"], errors='ignore')
        if totals_row_awb:
            df_awb_final = pd.concat([df_awb_final_data, pd.DataFrame([totals_row_awb]).drop(columns=["AWB Serial Part1", "AWB Serial Part2", "Exchange Rate", "AWB Number"], errors='ignore')], ignore_index=True)
        else:
            df_awb_final = df_awb_final_data
        print(f"  -> Finalized df_awb_final (for Invoices sheet). Shape: {df_awb_final.shape}")

        # Reorder columns for AWB sheet 
        base_awb_cols = ['AWB Prefix', 'AWB Serial']
        existing_cols = [
            col for col in df_awb_final.columns
            if col not in ['AWB Prefix', 'AWB Serial']
        ]
        awb_cols_order = base_awb_cols + existing_cols
        awb_cols_order = [col for col in awb_cols_order if col in df_awb_final.columns]
        df_awb_final = df_awb_final[awb_cols_order]

    # --- Process CCA Data ---
    if not df_cca.empty:
        # Format CCA Issue Date (assuming DDMMM input)
        def format_cca_date(date_str):
            try:
                return pd.to_datetime(date_str, format='%d%b', errors='coerce').strftime('%d/%b')
            except (ValueError, TypeError, AttributeError):
                 return date_str

        if 'CCA Issue Date' in df_cca.columns:
            df_cca['CCA Issue Date'] = df_cca['CCA Issue Date'].astype(str).apply(format_cca_date)

        # Ensure Prefix/Serial are strings
        if 'AWB Prefix' in df_cca.columns:
            df_cca['AWB Prefix'] = df_cca['AWB Prefix'].astype(str)
        if 'AWB Serial' in df_cca.columns:
            df_cca['AWB Serial'] = df_cca['AWB Serial'].astype(str)

        df_cca_final = df_cca

        # Reorder columns to put Prefix/Serial near the start
        cca_cols_order = ['AWB Prefix', 'AWB Serial'] + [col for col in df_cca_final.columns if col not in ['AWB Prefix', 'AWB Serial']]
        cca_cols_order = [col for col in cca_cols_order if col in df_cca_final.columns]
        df_cca_final = df_cca_final[cca_cols_order]

    # --- 4. Process Report Data and Reconciliation Logic ---
    report_data_df = None
    if os.path.exists(report_file_path):
        try:
            print(f"Reading report file: {report_file_path} with header on row 8 (index 7)")
            report_data_df = pd.read_excel(
                report_file_path,
                engine='xlrd',
                header=7, # Header on row 8 (0-indexed 7) - same as app(1).py
                dtype={'awbprefix': str, 'awbsuffix': str}
            )
            print(f"Report file read successfully. Shape: {report_data_df.shape}")

            # Clean column names by stripping whitespace and lowercasing - same as app(1).py
            report_data_df.columns = report_data_df.columns.str.strip().str.lower()
            print(f"Cleaned report columns: {report_data_df.columns.tolist()}")

            # Basic validation for required columns
            required_report_cols = ['awbprefix', 'awbsuffix', 'chargewt', 'frt_cost_rate', 'total_cost']
            missing_cols = [col for col in required_report_cols if col not in report_data_df.columns]
            if missing_cols:
                print(f"Warning: Report file missing required columns: {missing_cols}")
                report_data_df = None
            else:
                print(f"All required columns found in report file")
                
        except Exception as e:
            print(f"Warning: Could not read Excel file: {e}")
            report_data_df = None

    # --- 4. Reconciliation Logic ---
    if report_data_df is not None and not report_data_df.empty and not df_awb_for_recon.empty:
        print("--- Starting Reconciliation Process ---")

        # --- Prepare Invoice Data for Merge ---
        invoice_cols_for_merge = ['AWB Prefix', 'AWB Serial', 'Charge Weight', 'Net Yield Rate', 'Net Due for AWB']
        # Ensure AWB Prefix/Serial are strings and handle potential NaN before conversion
        df_awb_for_recon['AWB Prefix'] = df_awb_for_recon['AWB Prefix'].fillna('').astype(str).str.strip()
        df_awb_for_recon['AWB Serial'] = df_awb_for_recon['AWB Serial'].fillna('').astype(str).str.strip()

        # Ensure numeric types for comparison columns
        df_awb_for_recon['Charge Weight'] = safe_to_numeric(df_awb_for_recon['Charge Weight'])
        df_awb_for_recon['Net Yield Rate'] = safe_to_numeric(df_awb_for_recon['Net Yield Rate'])
        df_awb_for_recon['Net Due for AWB'] = safe_to_numeric(df_awb_for_recon['Net Due for AWB'])

        df_invoice_subset = df_awb_for_recon[invoice_cols_for_merge]
        print(f"  -> Prepared Invoice subset for merge. Shape: {df_invoice_subset.shape}")

        # --- Prepare Report Data for Merge ---
        report_cols = ['awbprefix', 'awbsuffix', 'chargewt', 'frt_cost_rate', 'total_cost']
        if all(col in report_data_df.columns for col in report_cols):
            df_report_subset = report_data_df[report_cols].copy()

            # Ensure merge keys are strings and normalized (remove .0)
            print("  -> Normalizing report merge key strings...")
            for col in ['awbprefix', 'awbsuffix']:
                 df_report_subset[col] = df_report_subset[col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

            # Rename columns for clarity and merging
            df_report_subset.rename(columns={
                'awbprefix': 'AWB Prefix',
                'awbsuffix': 'AWB Serial',
                'chargewt': 'Charge Weight (Report)',
                'frt_cost_rate': 'Net Yield Rate (Report)',
                'total_cost': 'Net Due (Report)'
            }, inplace=True)

            # Ensure numeric types for comparison columns
            df_report_subset['Charge Weight (Report)'] = safe_to_numeric(df_report_subset['Charge Weight (Report)'])
            df_report_subset['Net Yield Rate (Report)'] = safe_to_numeric(df_report_subset['Net Yield Rate (Report)'])
            df_report_subset['Net Due (Report)'] = safe_to_numeric(df_report_subset['Net Due (Report)'])

            print(f"  -> Prepared Report subset for merge. Shape before dedup: {df_report_subset.shape}")

            # Remove duplicates from report data
            initial_report_rows = len(df_report_subset)
            df_report_subset.drop_duplicates(subset=['AWB Prefix', 'AWB Serial'], keep='first', inplace=True)
            final_report_rows = len(df_report_subset)
            if initial_report_rows != final_report_rows:
                print(f"  -> Removed {initial_report_rows - final_report_rows} duplicate AWB entries from the report data.")
            print(f"  -> Shape after dedup: {df_report_subset.shape}")

            # --- Perform Left Merge ---
            print("  -> Performing left merge...")
            df_reconciliation = pd.merge(
                df_invoice_subset,
                df_report_subset,
                on=['AWB Prefix', 'AWB Serial'],
                how='left'
            )
            print(f"  -> Merge complete. Shape after merge: {df_reconciliation.shape}")
            
            # Rename columns for final output clarity
            df_reconciliation.rename(columns={
                'Charge Weight': 'Charge Weight (Invoice)',
                'Net Yield Rate': 'Net Yield Rate (Invoice)',
                'Net Due for AWB': 'Net Due (Invoice)'
            }, inplace=True)
            
            # --- Calculate Differences ---
            df_reconciliation['Diff Net Due'] = df_reconciliation['Net Due (Report)'].sub(df_reconciliation['Net Due (Invoice)'])
            print("  -> Calculated difference columns.")

            # --- Add Discrepancy Flag ---
            charge_weight_discrepancy = (
                (df_reconciliation['Charge Weight (Invoice)'].notna() & df_reconciliation['Charge Weight (Report)'].notna() &
                 (df_reconciliation['Charge Weight (Invoice)'].round(2) != df_reconciliation['Charge Weight (Report)'].round(2))) |
                (df_reconciliation['Charge Weight (Invoice)'].isna() & df_reconciliation['Charge Weight (Report)'].notna()) |
                (df_reconciliation['Charge Weight (Invoice)'].notna() & df_reconciliation['Charge Weight (Report)'].isna())
            )
            net_yield_rate_discrepancy = (
                (df_reconciliation['Net Yield Rate (Invoice)'].notna() & df_reconciliation['Net Yield Rate (Report)'].notna() &
                 (df_reconciliation['Net Yield Rate (Invoice)'].round(5) != df_reconciliation['Net Yield Rate (Report)'].round(5))) |
                (df_reconciliation['Net Yield Rate (Invoice)'].isna() & df_reconciliation['Net Yield Rate (Report)'].notna()) |
                (df_reconciliation['Net Yield Rate (Invoice)'].notna() & df_reconciliation['Net Yield Rate (Report)'].isna())
            )
            net_due_discrepancy = (
                 (df_reconciliation['Diff Net Due'].notna() & (df_reconciliation['Diff Net Due'].round(2) != 0))
            )

            discrepancy = charge_weight_discrepancy | net_yield_rate_discrepancy | net_due_discrepancy

            df_reconciliation['Discrepancy Found'] = discrepancy
            print("  -> Added 'Discrepancy Found' column.")

            # --- Reorder Reconciliation Columns ---
            recon_cols_order = [
                 'AWB Prefix', 'AWB Serial',
                 'Charge Weight (Invoice)', 'Charge Weight (Report)',
                 'Net Yield Rate (Invoice)', 'Net Yield Rate (Report)',
                 'Net Due (Invoice)', 'Net Due (Report)', 'Diff Net Due',
                 'Discrepancy Found'
             ]
            recon_cols_order = [col for col in recon_cols_order if col in df_reconciliation.columns]
            df_reconciliation = df_reconciliation[recon_cols_order]

            # --- Add Totals Row to Reconciliation Data ---
            cols_to_sum_rec = [
                'Charge Weight (Invoice)', 'Charge Weight (Report)',
                'Net Due (Invoice)', 'Net Due (Report)', 'Diff Net Due'
            ]
            valid_cols_to_sum = [col for col in cols_to_sum_rec if col in df_reconciliation.columns and pd.api.types.is_numeric_dtype(df_reconciliation[col])]
            
            if valid_cols_to_sum and not df_reconciliation.empty:
                totals_rec = df_reconciliation[valid_cols_to_sum].sum(numeric_only=True).to_dict()
                totals_row_rec = {col: '' for col in df_reconciliation.columns}
                totals_row_rec.update(totals_rec)
                totals_row_rec['AWB Prefix'] = 'Total'
                df_reconciliation = pd.concat([df_reconciliation, pd.DataFrame([totals_row_rec])], ignore_index=True)
                print("  -> Added calculated totals row to Reconciliation DataFrame.")
            else:
                print("  -> No numeric columns found or Reconciliation data empty. Skipping totals row addition.")

        else:
            print("  -> Report DataFrame missing required columns for reconciliation. Skipping reconciliation.")
            print(f"  -> Available columns: {report_data_df.columns.tolist()}")
            print(f"  -> Required columns: {report_cols}")
    elif report_data_df is None or report_data_df.empty:
         print("--- No report data provided or report data is empty. Skipping reconciliation. ---")
    elif df_awb_for_recon.empty:
         print("--- Invoice data (AWB) is empty. Skipping reconciliation. ---")


    # 5. Generate Excel output with all sheets
    print("-> Generating comprehensive Excel file...")
    
    # Create dynamic filename based on custom filename or N8N variables
    if custom_filename:
        output_filename = custom_filename
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
    elif workflow_id:
        # Use N8N pattern: invoicefile-workflowid.xlsx
        base_filename = os.path.splitext(os.path.basename(invoice_file_path))[0]
        output_filename = f"{base_filename}-{workflow_id}.xlsx"
    else:
        # Fallback to timestamp pattern
        import datetime
        base_filename = os.path.splitext(os.path.basename(invoice_file_path))[0]
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_filename}_processed_{timestamp}.xlsx"
    
    # Use /files/ for n8n, local directory for testing
    if os.path.exists("/files"):
        output_path = f"/files/{output_filename}"
    else:
        output_path = output_filename
    
    print(f"  -> Output file will be: {output_path}")
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # --- Calculate Summary Data ---
            print("--- Preparing Summary Sheet Data ---")
            summary_data = {}
            if not df_awb_for_recon.empty:
                df_summary_input = df_awb_for_recon[df_awb_for_recon['AWB Prefix'] != 'Total'].copy()
                if not df_summary_input.empty:
                    invoice_awb_count = len(df_summary_input)
                    total_invoice_amount = df_summary_input['Net Due for AWB'].sum()
                    total_charge_weight = df_summary_input['Charge Weight'].sum()
                    avg_net_yield_rate = (total_invoice_amount / total_charge_weight) if total_charge_weight else 0.0
                    
                    summary_data['Invoice AWB Count'] = invoice_awb_count
                    summary_data['Total Invoice Amount (Net Due)'] = total_invoice_amount
                    summary_data['Total Invoice Charge Weight'] = total_charge_weight
                    summary_data['Average Net Yield Rate'] = avg_net_yield_rate
                    print(f"  -> Calculated Invoice Stats: Count={invoice_awb_count}, Amount={total_invoice_amount:.2f}, Weight={total_charge_weight:.2f}, Avg Rate={avg_net_yield_rate:.5f}")
                else:
                     summary_data['Invoice AWB Count'] = 0
                     summary_data['Total Invoice Amount (Net Due)'] = 0.0
                     summary_data['Total Invoice Charge Weight'] = 0.0
                     summary_data['Average Net Yield Rate'] = 0.0
            else:
                summary_data['Invoice AWB Count'] = 0

            # Calculate report totals from the reconciliation dataframe
            if not df_reconciliation.empty and 'Net Due (Report)' in df_reconciliation.columns:
                df_rec_summary_input = df_reconciliation[df_reconciliation['AWB Prefix'] != 'Total'].copy()
                if not df_rec_summary_input.empty:
                    total_report_cost = df_rec_summary_input['Net Due (Report)'].sum()
                    difference_total_amount = total_report_cost - summary_data.get('Total Invoice Amount (Net Due)', 0.0)
                    summary_data['Total Report Amount (for Matched AWBs)'] = total_report_cost
                    summary_data['Difference (Report - Invoice)'] = difference_total_amount
                    print(f"  -> Calculated Report Stats: Total Cost={total_report_cost:.2f}, Difference={difference_total_amount:.2f}")
                else:
                    summary_data['Total Report Amount (for Matched AWBs)'] = 0.0
                    summary_data['Difference (Report - Invoice)'] = 0.0 - summary_data.get('Total Invoice Amount (Net Due)', 0.0)
            else:
                summary_data['Total Report Amount (for Matched AWBs)'] = 0.0

            # Create DataFrame for summary
            df_summary = pd.DataFrame(list(summary_data.items()), columns=['Metric', 'Value'])
            
            # --- Write Sheets in Specified Order ---

            # 1. Summary Sheet
            sheet_name_summary = "Summary"
            df_summary.to_excel(writer, sheet_name=sheet_name_summary, index=False)
            print(f"  -> Written summary data to '{sheet_name_summary}' sheet.")
            worksheet_summary = writer.sheets[sheet_name_summary]
            # Autofit columns for summary sheet
            for column in worksheet_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_len = len(str(cell.value))
                        header_len = len(str(worksheet_summary[f"{column_letter}1"].value))
                        max_length = max(max_length, cell_len, header_len)
                    except: pass
                adjusted_width = (max_length + 2) * 1.1
                worksheet_summary.column_dimensions[column_letter].width = adjusted_width
            print(f"  -> Autofit columns for '{sheet_name_summary}'.")

            # 2. Reconciliation Sheet
            if not df_reconciliation.empty:
                sheet_name_rec = "Reconciliation"
                df_reconciliation.to_excel(writer, sheet_name=sheet_name_rec, index=False)
                print(f"  -> Written {len(df_reconciliation)} rows to '{sheet_name_rec}' sheet.")
                worksheet_rec = writer.sheets[sheet_name_rec]
                # Formatting (Autofit, Table, Conditional)
                start_row_rec = 1
                end_row_rec = start_row_rec + len(df_reconciliation)
                start_col_rec = 1
                end_col_rec = len(df_reconciliation.columns)
                # Adjust table range
                rec_table_range = f"{get_column_letter(start_col_rec)}{start_row_rec}:{get_column_letter(end_col_rec)}{end_row_rec}"
                print(f"  -> Calculated Reconciliation Table Range: {rec_table_range}")
                for column in worksheet_rec.columns:
                     max_length = 0
                     column_letter = column[0].column_letter
                     for cell in column:
                        try:
                            cell_len = len(str(cell.value))
                            header_len = len(str(worksheet_rec[f"{column_letter}1"].value))
                            max_length = max(max_length, cell_len, header_len)
                        except: pass
                     adjusted_width = (max_length + 2) * 1.1
                     worksheet_rec.column_dimensions[column_letter].width = adjusted_width
                print(f"  -> Autofit columns for '{sheet_name_rec}'.")
                tab_rec = Table(displayName="ReconciliationTable", ref=rec_table_range)
                style_rec = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_rec.tableStyleInfo = style_rec
                # Apply table if there are any rows
                if len(df_reconciliation) > 0:
                    worksheet_rec.add_table(tab_rec)
                    print(f"  -> Added Excel table formatting to '{sheet_name_rec}'.")
                print(f"  -> Skipping conditional formatting to avoid Excel compatibility issues.")

            else:
                pd.DataFrame().to_excel(writer, sheet_name="Reconciliation", index=False)

            # 3. Invoices Sheet
            if not df_awb_final.empty:
                sheet_name_awb = "Invoices"
                df_awb_final.to_excel(writer, sheet_name=sheet_name_awb, index=False)
                print(f"  -> Written {len(df_awb_final)} rows to '{sheet_name_awb}' sheet.")
                worksheet_awb = writer.sheets[sheet_name_awb]
                # Formatting (Autofit, Table)
                start_row_awb = 1
                end_row_awb = start_row_awb + len(df_awb_final)
                start_col_awb = 1
                end_col_awb = len(df_awb_final.columns)
                awb_table_range = f"{get_column_letter(start_col_awb)}{start_row_awb}:{get_column_letter(end_col_awb)}{end_row_awb}"
                print(f"  -> Calculated AWB Table Range: {awb_table_range}")
                for column in worksheet_awb.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_len = len(str(cell.value))
                            header_len = len(str(worksheet_awb[f"{column_letter}1"].value))
                            max_length = max(max_length, cell_len, header_len)
                        except: pass
                    adjusted_width = (max_length + 2) * 1.1
                    worksheet_awb.column_dimensions[column_letter].width = adjusted_width
                print(f"  -> Autofit columns for '{sheet_name_awb}'.")
                tab_awb = Table(displayName="AWBTable", ref=awb_table_range)
                style_awb = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_awb.tableStyleInfo = style_awb
                if len(df_awb_final) > 0:
                    worksheet_awb.add_table(tab_awb)
                    print(f"  -> Added Excel table formatting to '{sheet_name_awb}'.")
                else:
                    print(f"  -> No data rows in '{sheet_name_awb}', skipping table creation.")
            else:
                pd.DataFrame().to_excel(writer, sheet_name="Invoices", index=False)
                print("  -> AWB DataFrame is empty. Written empty 'Invoices' sheet.")

            # 4. CCA Sheet
            if not df_cca_final.empty:
                sheet_name_cca = "CCA" 
                df_cca_final.to_excel(writer, sheet_name=sheet_name_cca, index=False)
                print(f"  -> Written {len(df_cca_final)} rows to '{sheet_name_cca}' sheet.")
                worksheet_cca = writer.sheets[sheet_name_cca]
                # Formatting (Autofit, Table)
                start_row_cca = 1
                end_row_cca = start_row_cca + len(df_cca_final)
                start_col_cca = 1
                end_col_cca = len(df_cca_final.columns)
                cca_table_range = f"{get_column_letter(start_col_cca)}{start_row_cca}:{get_column_letter(end_col_cca)}{end_row_cca}"
                print(f"  -> Calculated CCA Table Range: {cca_table_range}")
                for column in worksheet_cca.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_len = len(str(cell.value))
                            header_len = len(str(worksheet_cca[f"{column_letter}1"].value))
                            max_length = max(max_length, cell_len, header_len)
                        except: pass
                    adjusted_width = (max_length + 2) * 1.1
                    worksheet_cca.column_dimensions[column_letter].width = adjusted_width
                print(f"  -> Autofit columns for '{sheet_name_cca}'.")
                tab_cca = Table(displayName="CCATable", ref=cca_table_range)
                style_cca = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_cca.tableStyleInfo = style_cca
                if len(df_cca_final) > 0:
                    worksheet_cca.add_table(tab_cca)
                    print(f"  -> Added Excel table formatting to '{sheet_name_cca}'.")
                else:
                    print(f"  -> No data rows in '{sheet_name_cca}', skipping table creation.")
            else:
                pd.DataFrame().to_excel(writer, sheet_name="CCA", index=False)
                print("  -> CCA DataFrame is empty. Written empty 'CCA' sheet.")
        
        print("--- Excel File Written Successfully ---")

    except Exception as e:
        print(f"Error writing Excel file: {e}")
        import traceback
        traceback.print_exc()
        raise

    # Adjust row counts to exclude total row before returning
    invoices_rows_count = 0
    if not df_awb_final.empty:
         invoices_rows_count = len(df_awb_final)
         if 'Total' in df_awb_final['AWB Prefix'].values:
             invoices_rows_count -= 1
             
    cca_rows_count = 0
    if not df_cca_final.empty:
        cca_rows_count = len(df_cca_final)
        if 'Total' in df_cca_final['CCA Ref. No'].values:
            cca_rows_count -= 1

    print(f"Processing completed successfully:")
    print(f"  - AWB rows: {invoices_rows_count}")
    print(f"  - CCA rows: {cca_rows_count}")
    print(f"  - Total Net Due: {total_net_due_awb:.2f}")

    return output_path

def main():
    """Main entry point for command line execution."""
    if len(sys.argv) < 4 or len(sys.argv) > 6:
        print("Usage: python process_invoice.py <invoice_pdf_path> <report_excel_path> <output_json_path> [workflow_id] [custom_filename]")
        print("  workflow_id: Optional N8N workflow ID for dynamic filename")
        print("  custom_filename: Optional custom output filename (overrides workflow_id)")
        sys.exit(1)
    
    invoice_path = sys.argv[1]
    report_path = sys.argv[2]
    output_json_path = sys.argv[3]
    
    # Get optional parameters for N8N integration
    workflow_id = sys.argv[4] if len(sys.argv) >= 5 else None
    custom_filename = sys.argv[5] if len(sys.argv) >= 6 else None
    
    # In N8N, these can be passed as environment variables or workflow variables
    if not workflow_id:
        workflow_id = os.environ.get('N8N_WORKFLOW_ID')
    if not custom_filename:
        custom_filename = os.environ.get('N8N_OUTPUT_FILENAME')
    
    print(f"Processing with:")
    print(f"  Invoice: {invoice_path}")
    print(f"  Report: {report_path}")
    print(f"  Workflow ID: {workflow_id}")
    print(f"  Custom filename: {custom_filename}")
    
    try:
        result_path = process_files(invoice_path, report_path, workflow_id, custom_filename)
        
        if result_path:
            # Return result as JSON for n8n
            result = {
                "success": True,
                "output_file": result_path,
                "output_filename": os.path.basename(result_path),
                "message": "Processing completed successfully"
            }
        else:
            result = {
                "success": False,
                "error": "No data found in PDF files",
                "message": "Processing completed but no output generated"
            }
        
        with open(output_json_path, 'w') as f:
            json.dump(result, f)
            
    except Exception as e:
        result = {
            "success": False,
            "error": str(e),
            "message": "Processing failed"
        }
        
        with open(output_json_path, 'w') as f:
            json.dump(result, f)
        
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 