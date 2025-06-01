import os
import re
import pandas as pd
import pdfplumber # Add pdfplumber import
from flask import Flask, request, render_template, send_from_directory, url_for, flash, redirect
from werkzeug.utils import secure_filename # Needed for secure file handling
# Import necessary openpyxl components for table formatting
from openpyxl.worksheet.table import Table, TableStyleInfo
# Import utility to convert column index to letter
from openpyxl.utils import get_column_letter
# Import openpyxl styles for conditional formatting
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
# The functions below are defined in this file, so the import is removed.
# from extract_tables import extract_awb_data, extract_cca_data

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = 'supersecretkey' # Needed for flash messages

ALLOWED_EXTENSIONS = {'pdf', 'docx'} # Keep original allowed types
ALLOWED_REPORT_EXTENSIONS = {'xls'} # For the report file

def extract_awb_data(pdf): # Changed signature to accept pdf object
    """
    Extracts AWB data from specific pages of a FlyDubai PDF invoice 
    by processing the extracted text lines with layout preservation.
    Handles the multi-line format (AWB line + Date/Rate line).
    Dynamically determines the end page based on CCA header.
    """
    # print(f"--- Starting PDF Text Extraction Process for: {pdf_path} ---") # pdf_path not available
    print(f"--- Starting AWB PDF Text Extraction Process ---")
    extracted_data = []
    # target_pages = range(1, 7) # Pages 2-7 (0-indexed: 1-6) # REMOVED HARDCODED RANGE

    # --- Regex Definitions ---
    # AWB Line (Copied from test_extraction.py - includes optional space before K)
    line1_regex = re.compile(
        r"^\s*141\s+(\d{7})\s+(\d)\s+TLV\s+([A-Z]{3})\s+"  
        r"(\d+\.\d{2}\s*K)\s+"                       # Charge Weight (e.g., 560.00K or 560.00 K) - Made space optional
        r"([\d\.\,-]+)\s+"                             # PP Freight Charge
        r"([\d\.\,-]+)\s+"                             # PP Due Airline
        r"([\d\.\,-]+)\s+"                             # CC Freight Charge
        r"([\d\.\,-]+)\s+"                             # CC Due Agent
        r"([\d\.\,-]+)\s+"                             # CC Due Airline
        r"([\d\.\,-]+)\s+"                             # Disc.
        r"([\d\.\,-]+)\s+"                             # Agency Comm.
        r"([\d\.\,-]+)\s+"                             # Taxes
        r"([\d\.\,-]+)\s+"                             # Others
        r"([\d\.\,-]+)\s+"                             # Net Due for AWB (1)
        r"(1\.00000000)\s+"                           # Exchange Rate
        r"I?\s*([\d\.\,-]+)\s*$"                        # Net Due for AWB (2) - Adjusted slightly based on test script logic
    )
    # Rate pattern (search within the second line)
    net_yield_rate_regex_search = re.compile(r"(\d+\.\d+)")    
    # --- End Regex Definitions ---

    all_lines = []
    
    # --- Determine Target Page Range Dynamically ---
    cca_start_page_index = -1
    num_pages = len(pdf.pages)
    print(f"Total pages in PDF: {num_pages}")
    # Start searching for CCA from page 2 (index 1) onwards
    for page_num in range(1, num_pages):
        try:
            page_to_check = pdf.pages[page_num]
            text_to_check = page_to_check.extract_text()
            if text_to_check and "Section B: CCA Details" in text_to_check:
                cca_start_page_index = page_num
                print(f"  -> Found 'Section B: CCA Details' header on page {page_num + 1}. AWB data ends before this.")
                break
        except Exception as e:
            print(f"Warning: Error checking page {page_num + 1} for CCA header: {e}")

    # If CCA header not found, process all pages from page 2 to the end
    if cca_start_page_index == -1:
        print("  -> 'Section B: CCA Details' not found. Processing AWB data until the end of the document.")
        end_page_index = num_pages
    else:
        end_page_index = cca_start_page_index

    target_pages = range(1, end_page_index) # Page range is 1 to end_page_index (exclusive)
    print(f"AWB Target Page Indices (0-based): {list(target_pages)}")
    # --- End Determine Target Page Range ---

    # --- Extract Text from Target Pages ---
    # try: # Remove outer try/except as PDF opening is handled outside
    # with pdfplumber.open(pdf_path) as pdf: # REMOVED - pdf object is passed in
    # print(f"PDF has {len(pdf.pages)} pages.")
    for page_num in target_pages:
        # Check if page_num is valid (should always be within num_pages now)
        # if page_num < len(pdf.pages):
        try:
            page = pdf.pages[page_num]
            print(f"Extracting text with layout from Page {page_num + 1}...")
            # Use same extraction settings as test script
            text = page.extract_text(x_tolerance=2, layout=True) 
            if text:
                page_lines = text.split('\n')
                print(f"  -> Extracted {len(page_lines)} lines from page {page_num + 1}.")
                all_lines.extend(page_lines)
            else:
                print(f"  -> No text extracted from page {page_num + 1}.")
        except Exception as e:
            print(f"Error extracting text from page {page_num + 1}: {e}")
            # Continue to next page if one fails
        # else:
        #     print(f"Warning: Page {page_num + 1} requested, but PDF only has {len(pdf.pages)} pages.") # Should not happen now

    # except pdfplumber.exceptions.PDFSyntaxError as pdf_err: # Handled outside
    #     print(f"Error reading PDF structure: {pdf_err}")
    #     return pd.DataFrame()
    # except Exception as e: # Handled outside
    #     print(f"An error occurred during PDF processing: {e}")
    #     import traceback
    #     traceback.print_exc()
    #     return pd.DataFrame() 
    # --- End Extract Text ---

    print(f"--- Total text lines collected from AWB pages: {len(all_lines)} ---")

    # --- Process the collected text lines --- (Copied from test_extraction.py)
    i = 0
    while i < len(all_lines):
        line1_raw = all_lines[i] 
        line1 = re.sub(r'\s{2,}', ' ', line1_raw).strip() 
        
        # Minimal debug print for app context
        # print(f"\n[Line Index {i}] Checking line (normalized): '{line1[:100]}...'") 

        match1 = line1_regex.match(line1)
        
        if match1:
            # print(f"\n[Line Index {i}] SUCCESS: AWB Line Matched!")
            
            flight_date = ""
            net_yield_rate = ""
            lines_consumed = 1 

            if i + 1 < len(all_lines):
                line2_raw = all_lines[i+1]
                line2 = re.sub(r'\s{2,}', ' ', line2_raw).strip()
                
                # print(f"  [Line Index {i+1}] Checking for Date/Rate: '{line2}'")
                
                parts = line2.split(' ')
                flight_date = "" 
                net_yield_rate = "" 

                if parts:
                     flight_date = parts[0] 
                     lines_consumed += 1
                     # print(f"    -> Date String Extracted: '{flight_date}'") 

                     rate_match = net_yield_rate_regex_search.search(line2)
                     if rate_match:
                         potential_rates = net_yield_rate_regex_search.findall(line2)
                         if len(potential_rates) > 0:
                             net_yield_rate = potential_rates[0]
                             # print(f"    -> Rate Found: {net_yield_rate}")
                         # else: print(f"    -> Rate Found via search, but not via findall on '{line2}'")
                     # else: print(f"    -> Rate Not Found (search failed on '{line2}')")
                else:
                     # print(f"    -> Date/Rate line '{line2}' produced no parts after splitting. Discarding AWB match from line {i}.")
                     flight_date = "" 
            # else: print(f"  [Line Index {i+1}] No subsequent line available for Date/Rate check.")

            if match1 and flight_date:
                # print(f"  -> ADDING RECORD (AWB line {i}, Date/Rate line {i+1})")
                groups = match1.groups()
                # Data cleaning and assignment (Copied from test_extraction.py) 
                cleaned_groups = [g.replace(',', '.').strip() if isinstance(g, str) else g for g in groups]
                
                awb_number = f"141 {cleaned_groups[0]} {cleaned_groups[1]}"
                destination = cleaned_groups[2]
                charge_weight = cleaned_groups[3] # Already includes K
                pp_freight = cleaned_groups[4]
                pp_due_airline = cleaned_groups[5]
                cc_freight = cleaned_groups[6]
                cc_due_agent = cleaned_groups[7]
                cc_due_airline = cleaned_groups[8]
                disc = cleaned_groups[9]
                agency_comm = cleaned_groups[10]
                taxes = cleaned_groups[11]
                others = cleaned_groups[12]
                net_due_awb1 = cleaned_groups[13]
                exchange_rate = cleaned_groups[14]
                # net_due_awb2 = cleaned_groups[15] # Field 15 is the last one matched by updated regex

                # Basic validation (optional) - Comparing the two Net Due values if regex captured both
                # if len(cleaned_groups) > 15 and net_due_awb1 != cleaned_groups[15]:
                #     print(f"Warning: Net Due mismatch for AWB {awb_number}. Line1: {line1} -> {net_due_awb1} vs {cleaned_groups[15]}")

                extracted_data.append({
                    "AWB Number": awb_number,
                    "AWB Serial Part1": cleaned_groups[0],
                    "AWB Serial Part2": cleaned_groups[1],
                    "Flight Date": flight_date,
                    "Origin": "TLV",
                    "Destination": destination,
                    "Charge Weight": charge_weight,
                    "Net Yield Rate": net_yield_rate,
                    "PP Freight Charge": pp_freight,
                    "PP Due Airline": pp_due_airline,
                    "CC Freight Charge": cc_freight,
                    "CC Due Agent": cc_due_agent,
                    "CC Due Airline": cc_due_airline,
                    "Disc.": disc,
                    "Agency Comm.": agency_comm,
                    "Taxes": taxes,
                    "Others": others,
                    "Net Due for AWB": net_due_awb1,
                    "Exchange Rate": exchange_rate
                })
                
                # print(f"  -> Advancing index by {lines_consumed} lines to {i + lines_consumed}")
                i += lines_consumed
                continue 
            else:
                # print(f"  -> NOT ADDING RECORD (AWB line {i}, Date/Rate line {i+1} invalid or missing date)")
                i += 1 
        else:
            i += 1

    print("--- Finished Processing Text Lines ---")
    df = pd.DataFrame(extracted_data)

    return df

def extract_cca_data(pdf): # Changed signature to accept pdf object
    """
    Extracts CCA data from page 9 of a FlyDubai PDF invoice
    by processing the extracted text lines with layout preservation.
    Handles the multi-line format for CCA entries.
    Dynamically finds the CCA page.
    """
    # print(f"--- Starting CCA PDF Text Extraction Process for: {pdf_path} ---") # pdf_path not available
    print(f"--- Starting CCA PDF Text Extraction Process ---")
    extracted_data = []
    target_page = -1 # Initialize target page index
    raw_text_cca_page = "" # Renamed variable

    # --- Regex Definitions ---
    # CCA Line 1: Adjusted based on raw text spacing and formats
    # Now attempts to match the *full* line structure, including trailing columns,
    # to avoid matching the incomplete duplicated lines.
    # cca_line1_regex = re.compile( ... ) # Old line-based regex (commented out/removed)
    
    # Regex to find the Destination and CCA Issue Date on the second line
    # (These might still be useful if extracted separately from the block)
    destination_regex_search = re.compile(r"\b([A-Z]{3})\b")
    cca_date_regex_search = re.compile(r"\b(\d{2}[A-Z]{3})\b")

    # --- NEW Block Regex --- 
    # Matches the two-line CCA structure within a larger text block
    # Uses re.DOTALL to allow . to match newline characters
    cca_block_regex = re.compile(
        r"^\s*(\d{5,})\s+"          # CCA Ref No (1)
        r"(\d{3})\s*"             # AWB Prefix (2)
        r"(\d{7}\s\d{1})\s*"           # AWB Serial (3) - Captures "NNNNNNN N"
        r"([A-Z]{3})\s+"          # Origin (4)
        r"(\S+)\s+"               # MOP Freight (5)
        r"(\S+)\s+"               # MOP Other (6)
        r"([\d().,-]+)\s+"         # Freight Charge (7) - Handles parentheses
        r"([\d().,-]+)\s+"         # Due Airline (8) - Handles parentheses
        r"([\d().,-]+)\s+"         # Due Agent (9) - Handles parentheses
        r"([\d().,-]+)"            # Disc. (10) - Handles parentheses
        r".*?"                     # Non-greedy match until next numeric field (handles Disc 2nd part, spacing)
        r"([\d().,-]+)\s+"         # Agency Comm. (11) - Handles parentheses
        r"([\d().,-]+)\s+"         # Taxes (12) - Handles parentheses
        r"([\d().,-]+)\s+"         # Others (13) - Handles parentheses
        r"([\d().,-]+)\s+"         # Net Due (Sale) (14) - Handles parentheses
        r"\d\.\d{2}\s+"          # Exchange Rate (ignore)
        r"[\d().,-]+\s*?"          # Net Due (Invoice) (ignore) - Modified to handle parentheses
        # --- Second Line --- 
        r"\n\s*"                  # Match newline and start of next line
        r"(\d{2}[A-Z]{3})?"        # CCA Issue Date (15) - Optional capture
        r".*?"                     # Non-greedy match until destination
        r"\b([A-Z]{3})\b"          # Destination (16)"
        , re.MULTILINE | re.DOTALL # Multiline for ^, Dotall for newline matching
    )
    # --- End Regex Definitions ---

    all_lines = [] # Keep for compatibility? Maybe remove later.
    raw_text_page_9 = ""
    # try: # Remove outer try/except
    #     with pdfplumber.open(pdf_path) as pdf: # REMOVED - pdf object is passed in
    num_pages = len(pdf.pages)
    print(f"PDF has {num_pages} pages (in CCA function).")
    
    # --- Find the CCA Page Dynamically --- 
    # Start searching after potential AWB pages (e.g., page 2 onwards, index 1)
    # Adjust start_search_page if AWB data could theoretically be on very few pages
    start_search_page = 1 
    print(f"Searching for 'Section B: CCA Details' starting from page {start_search_page + 1}...")
    for page_num in range(start_search_page, num_pages):
        print(f"  Checking page {page_num + 1}...")
        try:
            page_to_check = pdf.pages[page_num]
            text_to_check = page_to_check.extract_text()
            if text_to_check and "Section B: CCA Details" in text_to_check:
                target_page = page_num
                print(f"  -> Found CCA header on page {target_page + 1}!")
                break # Stop searching once found
        except Exception as e:
            print(f"Warning: Error checking page {page_num + 1} for CCA header: {e}")
    # --- End Find Page ---

    if target_page != -1: # Check if CCA page was found
        try:
            page = pdf.pages[target_page]
            # --- Use standard extraction --- 
            print(f"Using standard text extraction (no layout=True) for Page {target_page + 1}.")
            raw_text_cca_page = page.extract_text()
            # --- 
            if raw_text_cca_page:
                print(f"  -> Extracted {len(raw_text_cca_page)} characters from page {target_page + 1}.")
            else:
                print(f"  -> No text extracted from page {target_page + 1} using standard extraction.")
        except Exception as e:
             print(f"Error extracting text from CCA page {target_page + 1}: {e}")
             raw_text_cca_page = "" # Ensure it's empty on error
    else:
        print("Warning: 'Section B: CCA Details' header not found in the document. Assuming no CCA data.")
        return pd.DataFrame() # Return empty DF if no CCA page found

    # --- Process using findall on the raw text block --- 
    print("--- Starting CCA processing using findall on raw text --- ")
    extracted_data = [] 
    
    # --- Define clean_currency locally ---
    def clean_currency(value_str):
        if isinstance(value_str, str):
            cleaned = value_str.replace(',', '').strip() # Remove thousand separators if any (CCA unlikely)
            is_negative = False
            
            # Scenario 1: Properly parenthesized e.g., "(123.45)"
            if cleaned.startswith('(') and cleaned.endswith(')'):
                is_negative = True
                cleaned = cleaned[1:-1] # Remove both parentheses
            # Scenario 2: Starts with '(', but no closing ')' from regex e.g., "(123.45"
            elif cleaned.startswith('(') and not cleaned.endswith(')'):
                # Try to strip the opening '(' and see if the rest is a number
                temp_cleaned = cleaned[1:]
                try:
                    # Ensure what remains is a potentially valid number string before float conversion attempt
                    # This regex check is basic, float() will do the final validation
                    if re.match(r"^[\d.]+$", temp_cleaned):
                        is_negative = True
                        cleaned = temp_cleaned # Use the content inside the opening parenthesis
                    # If not a simple number after stripping '(', let it fall through to normal float conversion / error
                except re.error: # In case of bad regex, though unlikely here
                    pass # Fall through

            try:
                value = float(cleaned)
                if is_negative:
                    value = -value
                return value
            except ValueError:
                # If conversion fails, return the cleaned string (or original if cleaning was minimal)
                # This handles cases like "PP", "CC", or unexpected text.
                if not cleaned and value_str == "()": # If original was "()" and now empty
                    return 0.0 # Or handle as None or raise an error based on requirements
                return value_str # Return original if it's not a number like "PP"
        return value_str
    # --- End clean_currency definition ---
    
    if raw_text_cca_page:
        matches = cca_block_regex.findall(raw_text_cca_page)
        print(f"  -> Found {len(matches)} potential CCA blocks using findall.")

        for groups in matches:
            # --- DEBUG: Print matched groups from findall (REMOVE) ---
            # print(f"  -> Processing Match Groups: {groups}")
            # --- END DEBUG ---
            
            # Check if groups has expected length (adjust based on final regex captures)
            if len(groups) == 16:
                 extracted_data.append({
                    "CCA Ref. No": groups[0],
                    "AWB Prefix": groups[1],
                    "AWB Serial": groups[2].replace(" ", ""), # Remove space from AWB Serial
                    "CCA Issue Date": groups[14] if groups[14] else "", # Extracted from block
                    "Origin": groups[3],
                    "Destination": groups[15], # Extracted from block
                    "MOP Freight Charge": groups[4],
                    "MOP Other Charge": groups[5],
                    "Freight Charge": clean_currency(groups[6]),
                    "Due Airline": clean_currency(groups[7]),
                    "Due Agent": clean_currency(groups[8]),
                    "Disc.": clean_currency(groups[9]),
                    "Agency Comm.": clean_currency(groups[10]),
                    "Taxes": clean_currency(groups[11]),
                    "Others": clean_currency(groups[12]),
                    "Net Due for AWB (Sale Currency)": clean_currency(groups[13]),
                })
                 # print("    -> Successfully appended record from match.") # REMOVE DEBUG
            else:
                 print(f"    -> WARNING: Match found but had unexpected number of groups ({len(groups)}). Skipping.")
                 # print(f"     -> Groups: {groups}") # Optional: print groups for debugging

    else:
        print("  -> No raw text extracted to process.")
    
    # --- End Restore ---

    # --- Process the collected CCA text lines --- (OLD CODE - REMOVE/IGNORE)
    # extracted_data = [] # Ensure extracted_data is initialized here
    # i = 0
    # while i < len(all_lines):
    # ... (rest of the old line processing loop) ... 
    #         i += 1
    
    print("--- Finished Processing CCA Text ---") # Updated message
    df_cca = pd.DataFrame(extracted_data)
    
    # Identify numeric columns for potential totaling (excluding AWB parts)
    numeric_cols_cca = [col for col in df_cca.columns if df_cca[col].apply(lambda x: isinstance(x, (int, float))).all()]
    
    # Add Totals Row if data exists and numeric columns are found
    if not df_cca.empty and numeric_cols_cca:
        totals_cca = df_cca[numeric_cols_cca].sum().to_dict()
        totals_row_cca = {col: totals_cca.get(col, '') for col in df_cca.columns}
        # Add a label for the totals row, e.g., in the 'CCA Ref. No' column
        totals_row_cca['CCA Ref. No'] = 'Total' 
        # Ensure AWB details aren't summed
        totals_row_cca['AWB Prefix'] = ''
        totals_row_cca['AWB Serial'] = ''
        totals_row_cca['CCA Issue Date'] = ''
        totals_row_cca['Origin'] = ''
        totals_row_cca['Destination'] = ''
        totals_row_cca['MOP Freight Charge'] = ''
        totals_row_cca['MOP Other Charge'] = ''
        
        df_cca = pd.concat([df_cca, pd.DataFrame([totals_row_cca])], ignore_index=True)
        print("  -> Added totals row to CCA DataFrame.")
    elif not df_cca.empty:
         print("  -> CCA DataFrame created, but no purely numeric columns found for totaling.")
    
    return df_cca


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/uploads/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/test')
def test():
    """Test route to process the sample PDF file directly and show results"""
    try:
        # Use the PDF sample file name
        sample_file_name = "2503013781418TLV010936_25-25.pdf" 
        sample_file_path = os.path.join(app.config['UPLOAD_FOLDER'], sample_file_name)
        
        if not os.path.exists(sample_file_path):
             # If not in uploads, maybe it's in the root? (Less likely now)
             script_dir = os.path.dirname(os.path.abspath(__file__))
             if os.path.exists(os.path.join(script_dir, sample_file_name)):
                 sample_file_path = os.path.join(script_dir, sample_file_name)
             else:
                 flash(f"Error: Sample file '{sample_file_name}' not found. Searched in '{app.config['UPLOAD_FOLDER']}' and script directory.")
                 return redirect(url_for('index')) # Redirect back

        results = process_file(sample_file_path) # process_file uses the new extract_awb_data
        
        if results["invoices_rows"] == 0: # Only check invoice rows now
             flash(f"Processing Complete: No AWB data found in '{sample_file_name}'. Excel file not generated.")
             return redirect(url_for('index'))

        # Render results on the index page using flash messages or render a new template
        flash(f"Test Processing Successful for '{sample_file_name}'")
        flash(f"AWB Rows Extracted: {results['invoices_rows']}")
        flash(f"CCA Rows Extracted: {results['cca_rows']}") # Add CCA count
        flash(f"Total Net Due for AWB: {results['total_net_due_awb']:.2f}")
        flash(f"<a href='{results['download_url']}'>Download Processed Excel File</a>")
        return redirect(url_for('index'))

    except Exception as e:
        error_message = f"Error during test processing: {e}"
        print(error_message)
        import traceback
        traceback.print_exc() # Print stack trace to console for debugging
        flash(f"<h1>Error processing test file</h1><p>{error_message}</p>", 'error')
        return redirect(url_for('index'))

# --- Function to safely convert to numeric, returning NaN on failure ---
def safe_to_numeric(series):
    return pd.to_numeric(series, errors='coerce')

def allowed_file(filename, allowed_set):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_set

def process_file(file_path, report_data_df=None): # Add report_data_df parameter
    """Process a specific PDF/DOCX file, merge with report data, and return the results"""
    # Ensure the file exists before processing
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return {
            "invoices_rows": 0,
            "cca_rows": 0,
            "excel_file": "",
            "download_url": "",
            "total_net_due_awb": 0.0
        }

    # 1. Extract AWB Data
    print(f"Calling extract_awb_data for: {file_path}")
    # df_awb = extract_awb_data(file_path) # OLD CALL
    
    # --- Separate Data from Potential Totals EARLY ---
    # Assuming extract_awb_data returns ONLY data rows
    # df_awb_data_only = df_awb.copy() # Will be assigned later
    # print(f"  -> Initial AWB data rows extracted: {len(df_awb_data_only)}") # Moved down

    # 2. Extract CCA Data
    print(f"Calling extract_cca_data for: {file_path}")
    # df_cca = extract_cca_data(file_path) # Call the new function # OLD CALL

    # --- Open PDF once ---
    df_awb = pd.DataFrame()
    df_cca = pd.DataFrame()
    try:
        with pdfplumber.open(file_path) as pdf:
            df_awb = extract_awb_data(pdf) # Pass pdf object
            df_cca = extract_cca_data(pdf) # Pass pdf object
    except pdfplumber.exceptions.PDFSyntaxError as pdf_err:
         print(f"Error reading PDF structure in process_file: {pdf_err}")
         # Return empty results if PDF is unreadable
         return {
            "invoices_rows": 0, "cca_rows": 0, "excel_file": "",
            "download_url": "", "total_net_due_awb": 0.0
         }
    except Exception as e:
         print(f"An error occurred during PDF processing in process_file: {e}")
         import traceback
         traceback.print_exc()
          # Return empty results on other errors
         return {
            "invoices_rows": 0, "cca_rows": 0, "excel_file": "",
            "download_url": "", "total_net_due_awb": 0.0
         }
    # --- End Open PDF ---

    df_awb_data_only = df_awb.copy() # Assign data after extraction
    print(f"  -> Initial AWB data rows extracted: {len(df_awb_data_only)}") # Print count here

    if df_awb.empty and df_cca.empty: # Check if both are empty
        print("Both extract_awb_data and extract_cca_data returned empty DataFrames.")
        return {
            "invoices_rows": 0,
            "cca_rows": 0,
            "excel_file": "",
            "download_url": "",
            "total_net_due_awb": 0.0
        }

    # Prepare Excel output filenames
    base_filename = os.path.splitext(os.path.basename(file_path))[0]
    excel_filename = f"{base_filename}_processed.xlsx"
    excel_filepath = os.path.join(app.config["UPLOAD_FOLDER"], excel_filename)
    download_url = url_for('download_file', filename=excel_filename)
    
    total_net_due_awb = 0.0 # Initialize total
    df_awb_final = pd.DataFrame() # Initialize df_awb_final (will include totals for its sheet)
    df_awb_for_recon = pd.DataFrame() # Initialize df_awb_for_recon (data ONLY)
    df_cca_final = pd.DataFrame() # Initialize df_cca_final
    df_reconciliation = pd.DataFrame() # Initialize df_reconciliation

    # --- Process AWB Data (on data_only first) ---
    if not df_awb_data_only.empty:
        # Process AWB data (calculations, column drops) before writing
        numeric_cols_awb = [
            'PP Freight Charge', 'PP Due Airline', 'CC Freight Charge', 'CC Due Agent',
            'CC Due Airline', 'Disc.', 'Agency Comm.', 'Taxes', 'Others', 'Net Due for AWB',
            'Net Yield Rate' # Add Net Yield Rate for conversion
        ]
        # Convert only potential numeric columns on the data_only DataFrame
        print("  -> Converting AWB columns to numeric (on data_only)...")
        for col in numeric_cols_awb:
             if col in df_awb_data_only.columns:
                # Use safe conversion
                df_awb_data_only[col] = safe_to_numeric(df_awb_data_only[col])

        # Calculate totals based *only* on data rows
        valid_numeric_cols_awb = [col for col in numeric_cols_awb if col in df_awb_data_only.columns and pd.api.types.is_numeric_dtype(df_awb_data_only[col])]
        cols_to_drop_for_sum = ["AWB Number", "Flight Date", "Origin", "Destination", "Charge Weight", "AWB Serial Part1", "AWB Serial Part2", "Exchange Rate"]
        df_numeric_awb_data_only = df_awb_data_only.drop(columns=[col for col in cols_to_drop_for_sum if col in df_awb_data_only.columns], errors='ignore')

        totals_row_awb = {} # Initialize totals row
        if valid_numeric_cols_awb and not df_numeric_awb_data_only.empty:
            totals_awb = df_numeric_awb_data_only[valid_numeric_cols_awb].sum(numeric_only=True).to_dict() # Ensure only numeric sum
            totals_row_awb = {col: totals_awb.get(col, '') for col in df_awb_data_only.columns} # Use data_only columns for structure
            totals_row_awb["AWB Number"] = "Total" # Label the totals row
            # Ensure non-numeric fields in total row are blank
            for col in cols_to_drop_for_sum:
                if col in totals_row_awb:
                    totals_row_awb[col] = ''
            
            # Calculate the overall total net due from the data_only df
            if 'Net Due for AWB' in df_awb_data_only.columns and pd.api.types.is_numeric_dtype(df_awb_data_only['Net Due for AWB']):
                total_net_due_awb = df_awb_data_only['Net Due for AWB'].sum() # Correct total based on data only
                print(f"  -> Calculated Total Net Due (data only): {total_net_due_awb}")
            else:
                total_net_due_awb = 0.0 # Fallback

            print("  -> Calculated totals row based on AWB data_only.")
        else:
            print("  -> AWB data_only empty or no valid numeric columns found for totaling.")
            if 'Net Due for AWB' in df_awb_data_only.columns and pd.api.types.is_numeric_dtype(df_awb_data_only['Net Due for AWB']):
               total_net_due_awb = df_awb_data_only['Net Due for AWB'].sum()

        # --- Apply Formatting and Splitting to df_awb_data_only --- 
        # Format Flight Date (assuming DDMMMYY input) on data_only
        def format_date(date_str):
            try:
                # Attempt parsing known format, be robust to errors
                return pd.to_datetime(date_str, format='%d%b%y', errors='coerce').strftime('%d/%m/%Y')
            except (ValueError, TypeError, AttributeError): # Added AttributeError and correct indentation
                return date_str # Return original if format is wrong or it's not a string

        if 'Flight Date' in df_awb_data_only.columns:
             # Ensure it's string before applying
             df_awb_data_only['Flight Date'] = df_awb_data_only['Flight Date'].astype(str).apply(format_date)

        # Format Charge Weight (remove ' K' and convert to numeric) on data_only
        if 'Charge Weight' in df_awb_data_only.columns:
            df_awb_data_only['Charge Weight'] = df_awb_data_only['Charge Weight'].astype(str).str.replace(r'\s*K$', '', regex=True)
            # Use safe conversion
            df_awb_data_only['Charge Weight'] = safe_to_numeric(df_awb_data_only['Charge Weight'])

        # -- Split AWB Number (for AWB Sheet and Reconciliation) on data_only --
        if "AWB Number" in df_awb_data_only.columns:
            # Split "141 ABCDEFG H" into "141" and "ABCDEFGH"
            split_awb = df_awb_data_only["AWB Number"] \
                .astype(str) \
                .str.split(r'\s+', n=2, expand=True)
            # Assign prefix, handling potential None from split
            df_awb_data_only['AWB Prefix'] = split_awb[0].fillna('').astype(str).str.strip()
            # Combine parts 1 and 2 for the serial, handling potential None
            df_awb_data_only['AWB Serial'] = split_awb[1].fillna('').astype(str) + split_awb[2].fillna('').astype(str)
            # Remove extra spaces within serial just in case
            df_awb_data_only['AWB Serial'] = df_awb_data_only['AWB Serial'].str.replace(r'\s+', '', regex=True).str.strip()
            # Update the separate totals_row_awb with split info if needed (optional, keeps structure consistent)
            if totals_row_awb:
                 totals_row_awb['AWB Prefix'] = 'Total'
                 totals_row_awb['AWB Serial'] = ''

        # --- Finalize DataFrames for Sheets ---
        # df_awb_for_recon is the processed data_only DataFrame
        df_awb_for_recon = df_awb_data_only.copy()
        print(f"  -> Finalized df_awb_for_recon (data only). Shape: {df_awb_for_recon.shape}")

        # df_awb_final (for Invoices sheet) = processed data_only + totals_row
        df_awb_final_data = df_awb_data_only.drop(columns=["AWB Serial Part1", "AWB Serial Part2", "Exchange Rate", "AWB Number"], errors='ignore')
        if totals_row_awb: # Only add total row if it was calculated
            df_awb_final = pd.concat([df_awb_final_data, pd.DataFrame([totals_row_awb]).drop(columns=["AWB Serial Part1", "AWB Serial Part2", "Exchange Rate", "AWB Number"], errors='ignore')], ignore_index=True)
        else:
            df_awb_final = df_awb_final_data
        print(f"  -> Finalized df_awb_final (for Invoices sheet). Shape: {df_awb_final.shape}")

        # Reorder columns for AWB sheet 
        base_awb_cols = ['AWB Prefix', 'AWB Serial']
        # Find existing cols from df_awb before merge, excluding merge keys and dropped cols
        existing_cols = [
            col for col in df_awb_final.columns
            if col not in ['AWB Prefix', 'AWB Serial']
        ]
        awb_cols_order = base_awb_cols + existing_cols
        # Ensure only columns actually present in the final df are selected
        awb_cols_order = [col for col in awb_cols_order if col in df_awb_final.columns]
        df_awb_final = df_awb_final[awb_cols_order]


    if not df_cca.empty:
        # CCA data is already processed (including totals) in extract_cca_data
        # Format CCA Issue Date (assuming DDMMM input)
        def format_cca_date(date_str):
            try:
                # Attempt parsing known format
                return pd.to_datetime(date_str, format='%d%b', errors='coerce').strftime('%d/%b') # Only day/month seems present
            except (ValueError, TypeError, AttributeError): # Added AttributeError
                 return date_str # Return original if format is wrong

        if 'CCA Issue Date' in df_cca.columns:
            # Ensure string before applying
            df_cca['CCA Issue Date'] = df_cca['CCA Issue Date'].astype(str).apply(format_cca_date)

        # Ensure Prefix/Serial are strings
        if 'AWB Prefix' in df_cca.columns:
            df_cca['AWB Prefix'] = df_cca['AWB Prefix'].astype(str)
        if 'AWB Serial' in df_cca.columns:
            df_cca['AWB Serial'] = df_cca['AWB Serial'].astype(str)

        # Use the df with separate Prefix/Serial
        df_cca_final = df_cca

        # Reorder columns to put Prefix/Serial near the start
        cca_cols_order = ['AWB Prefix', 'AWB Serial'] + [col for col in df_cca_final.columns if col not in ['AWB Prefix', 'AWB Serial']]
        # Filter to ensure all columns exist in the dataframe before reordering
        cca_cols_order = [col for col in cca_cols_order if col in df_cca_final.columns]
        df_cca_final = df_cca_final[cca_cols_order]

    # --- 4. Reconciliation Logic ---
    if report_data_df is not None and not report_data_df.empty and not df_awb_for_recon.empty:
        print("--- Starting Reconciliation Process ---")

        # --- Prepare Invoice Data for Merge (using df_awb_for_recon) ---
        # Select columns needed from df_awb_for_recon (guaranteed no totals row)
        invoice_cols_for_merge = ['AWB Prefix', 'AWB Serial', 'Charge Weight', 'Net Yield Rate', 'Net Due for AWB']
        # Ensure AWB Prefix/Serial are strings and handle potential NaN before conversion
        df_awb_for_recon['AWB Prefix'] = df_awb_for_recon['AWB Prefix'].fillna('').astype(str).str.strip()
        df_awb_for_recon['AWB Serial'] = df_awb_for_recon['AWB Serial'].fillna('').astype(str).str.strip()

        # Filtered data is already assigned to df_awb_for_recon
        # Ensure numeric types for comparison columns in df_awb_for_recon
        df_awb_for_recon['Charge Weight'] = safe_to_numeric(df_awb_for_recon['Charge Weight'])
        df_awb_for_recon['Net Yield Rate'] = safe_to_numeric(df_awb_for_recon['Net Yield Rate'])
        df_awb_for_recon['Net Due for AWB'] = safe_to_numeric(df_awb_for_recon['Net Due for AWB'])

        df_invoice_subset = df_awb_for_recon[invoice_cols_for_merge]
        print(f"  -> Prepared Invoice subset for merge from df_awb_for_recon. Shape: {df_invoice_subset.shape}")
        # print(f"  -> Invoice subset dtypes:\n{df_invoice_subset.dtypes}") # Debug

        # --- Prepare Report Data for Merge ---
        report_cols = ['awbprefix', 'awbsuffix', 'chargewt', 'frt_cost_rate', 'total_cost']
        # Column names in report_data_df are already lowercased and stripped in upload()
        if all(col in report_data_df.columns for col in report_cols):
            df_report_subset = report_data_df[report_cols].copy()

            # Ensure merge keys are strings and normalized (remove .0)
            print("  -> Normalizing report merge key strings...")
            for col in ['awbprefix', 'awbsuffix']:
                 df_report_subset[col] = df_report_subset[col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

            # Rename columns for clarity and merging
            df_report_subset.rename(columns={
                'awbprefix': 'AWB Prefix', # Match invoice name for merging
                'awbsuffix': 'AWB Serial', # Match invoice name for merging
                'chargewt': 'Charge Weight (Report)',
                'frt_cost_rate': 'Net Yield Rate (Report)',
                'total_cost': 'Net Due (Report)'
            }, inplace=True)

            # Ensure numeric types for comparison columns
            df_report_subset['Charge Weight (Report)'] = safe_to_numeric(df_report_subset['Charge Weight (Report)'])
            df_report_subset['Net Yield Rate (Report)'] = safe_to_numeric(df_report_subset['Net Yield Rate (Report)'])
            df_report_subset['Net Due (Report)'] = safe_to_numeric(df_report_subset['Net Due (Report)'])

            print(f"  -> Prepared Report subset for merge. Shape before dedup: {df_report_subset.shape}") # Modified print
            # print(f"  -> Report subset dtypes:\\n{df_report_subset.dtypes}") # Debug
            # print(f"  -> Report subset head:\\n{df_report_subset.head()}") # Debug

            # --- ADDED: Remove duplicates from report data based on AWB keys ---
            initial_report_rows = len(df_report_subset)
            df_report_subset.drop_duplicates(subset=['AWB Prefix', 'AWB Serial'], keep='first', inplace=True)
            final_report_rows = len(df_report_subset)
            if initial_report_rows != final_report_rows:
                print(f"  -> Removed {initial_report_rows - final_report_rows} duplicate AWB entries from the report data.")
            print(f"  -> Shape after dedup: {df_report_subset.shape}")
            # --- END ADDED ---


            # --- Perform Left Merge (Invoice-based Reconciliation) ---
            print("  -> Performing left merge...")
            df_reconciliation = pd.merge(
                df_invoice_subset,
                df_report_subset,
                on=['AWB Prefix', 'AWB Serial'],
                how='left', # Changed from 'outer' to 'left'
                suffixes=None # No suffix needed on left merge if column names are distinct before merge
            )
            print(f"  -> Merge complete. Shape after merge: {df_reconciliation.shape}")
            # print(f"  -> Merged columns: {df_reconciliation.columns.tolist()}") # Debug

            # Rename columns for final output clarity
            # Note: Charge Weight, Net Yield Rate, Net Due for AWB already have numeric types from prep
            df_reconciliation.rename(columns={
                'Charge Weight': 'Charge Weight (Invoice)',
                'Net Yield Rate': 'Net Yield Rate (Invoice)',
                'Net Due for AWB': 'Net Due (Invoice)'
                # Report columns were already renamed before merge
            }, inplace=True)

            # --- Calculate Differences ---
            # Use .sub() which handles NaNs gracefully (result is NaN if either input is NaN)
            # Keep only Net Due difference
            # df_reconciliation['Diff Charge Weight'] = df_reconciliation['Charge Weight (Report)'].sub(df_reconciliation['Charge Weight (Invoice)'])
            # df_reconciliation['Diff Net Yield Rate'] = df_reconciliation['Net Yield Rate (Report)'].sub(df_reconciliation['Net Yield Rate (Invoice)'])
            df_reconciliation['Diff Net Due'] = df_reconciliation['Net Due (Report)'].sub(df_reconciliation['Net Due (Invoice)'])
            print("  -> Calculated difference columns.")

            # --- Add Discrepancy Flag ---
            # Compare Charge Weight and Net Yield Rate directly, check Diff Net Due
            # Also check if values exist in one file but not the other (NaN comparison)
            # Use round() for float comparison to handle potential precision issues
            charge_weight_discrepancy = (
                (df_reconciliation['Charge Weight (Invoice)'].notna() & df_reconciliation['Charge Weight (Report)'].notna() &
                 (df_reconciliation['Charge Weight (Invoice)'].round(2) != df_reconciliation['Charge Weight (Report)'].round(2))) |
                (df_reconciliation['Charge Weight (Invoice)'].isna() & df_reconciliation['Charge Weight (Report)'].notna()) |
                (df_reconciliation['Charge Weight (Invoice)'].notna() & df_reconciliation['Charge Weight (Report)'].isna())
            )
            net_yield_rate_discrepancy = (
                (df_reconciliation['Net Yield Rate (Invoice)'].notna() & df_reconciliation['Net Yield Rate (Report)'].notna() &
                 (df_reconciliation['Net Yield Rate (Invoice)'].round(5) != df_reconciliation['Net Yield Rate (Report)'].round(5))) | # Using more precision for rates
                (df_reconciliation['Net Yield Rate (Invoice)'].isna() & df_reconciliation['Net Yield Rate (Report)'].notna()) |
                (df_reconciliation['Net Yield Rate (Invoice)'].notna() & df_reconciliation['Net Yield Rate (Report)'].isna())
            )
            net_due_discrepancy = (
                 (df_reconciliation['Diff Net Due'].notna() & (df_reconciliation['Diff Net Due'].round(2) != 0))
            )

            discrepancy = charge_weight_discrepancy | net_yield_rate_discrepancy | net_due_discrepancy

            df_reconciliation['Discrepancy Found'] = discrepancy
            print("  -> Added 'Discrepancy Found' column.")

            # --- Reorder Reconciliation Columns ---
            # Remove the direct diff columns for weight and rate
            recon_cols_order = [
                 'AWB Prefix', 'AWB Serial',
                 'Charge Weight (Invoice)', 'Charge Weight (Report)',
                 'Net Yield Rate (Invoice)', 'Net Yield Rate (Report)',
                 'Net Due (Invoice)', 'Net Due (Report)', 'Diff Net Due',
                 'Discrepancy Found'
             ]
             # Ensure all columns exist before reordering
            recon_cols_order = [col for col in recon_cols_order if col in df_reconciliation.columns]
            df_reconciliation = df_reconciliation[recon_cols_order]

            # --- Identify AWBs with Net Due Discrepancies for Invoice Sheet Highlighting ---
            discrepancy_awbs = set()
            # Filter out the potential total row and rows where Diff Net Due is NaN or zero (within tolerance)
            if 'Diff Net Due' in df_reconciliation.columns and 'AWB Prefix' in df_reconciliation.columns and 'AWB Serial' in df_reconciliation.columns:
                recon_data_only = df_reconciliation[df_reconciliation['AWB Prefix'] != 'Total'].copy()
                discrepancy_rows = recon_data_only[
                    recon_data_only['Diff Net Due'].notna() & 
                    (recon_data_only['Diff Net Due'].round(2) != 0)
                ]
                if not discrepancy_rows.empty:
                    discrepancy_awbs = set(zip(
                        discrepancy_rows['AWB Prefix'].astype(str),
                        discrepancy_rows['AWB Serial'].astype(str)
                    ))
                    print(f"  -> Identified {len(discrepancy_awbs)} AWBs with non-zero Net Due difference for invoice highlighting.")
                else:
                    print("  -> No AWBs found with non-zero Net Due difference in reconciliation data.")
            else:
                print("  -> Could not identify discrepancy AWBs due to missing columns in reconciliation data.")
            # --- End Identification ---

            # --- Add Totals Row to Reconciliation Data --- (RE-ADDING LOGIC)
            cols_to_sum_rec = [
                'Charge Weight (Invoice)', 'Charge Weight (Report)',
                'Net Due (Invoice)', 'Net Due (Report)', 'Diff Net Due'
            ]
            # Calculate totals only for numeric columns present in the DataFrame
            valid_cols_to_sum = [col for col in cols_to_sum_rec if col in df_reconciliation.columns and pd.api.types.is_numeric_dtype(df_reconciliation[col])]
            
            if valid_cols_to_sum and not df_reconciliation.empty: # Check df is not empty
                totals_rec = df_reconciliation[valid_cols_to_sum].sum(numeric_only=True).to_dict()
                totals_row_rec = {col: '' for col in df_reconciliation.columns} # Initialize with blanks
                totals_row_rec.update(totals_rec) # Add sums
                totals_row_rec['AWB Prefix'] = 'Total' # Add label
                # Append totals row DataFrame
                df_reconciliation = pd.concat([df_reconciliation, pd.DataFrame([totals_row_rec])], ignore_index=True)
                print("  -> Added calculated totals row to Reconciliation DataFrame.")
            else:
                print("  -> No numeric columns found or Reconciliation data empty. Skipping totals row addition.")

        else:
            print("  -> Report DataFrame missing required columns for reconciliation. Skipping reconciliation.")
            flash("Warning: Report file missing required columns (chargewt, frt_cost_rate, total_cost). Reconciliation skipped.", 'warning')
            # df_reconciliation remains empty
    elif report_data_df is None or report_data_df.empty:
         print("--- No report data provided or report data is empty. Skipping reconciliation. ---")
    elif df_awb_for_recon.empty:
         print("--- Invoice data (AWB) is empty. Skipping reconciliation. ---")


    # 5. Write to Excel using ExcelWriter for multiple sheets
    try:
        print(f"Writing data to Excel file: {excel_filepath}")
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:

            # --- Calculate Summary Data ---
            # (Calculation remains here, but depends on correctly filtered df_awb_filtered)
            print("--- Preparing Summary Sheet Data ---")
            summary_data = {}
            # Use df_awb_filtered which *should* exclude the total row
            # Re-apply filter for safety right before summing:
            if 'df_awb_for_recon' in locals() and not df_awb_for_recon.empty:
                # Ensure the filter is applied to the DF we are about to sum (using df_awb_for_recon now)
                df_summary_input = df_awb_for_recon[df_awb_for_recon['AWB Prefix'] != 'Total'].copy()
                if not df_summary_input.empty:
                    invoice_awb_count = len(df_summary_input)
                    # Use original column names from df_summary_input (already numeric)
                    total_invoice_amount = df_summary_input['Net Due for AWB'].sum()
                    total_charge_weight = df_summary_input['Charge Weight'].sum()
                    # Calculate Average Net Yield Rate = Total Net Due / Total Charge Weight
                    avg_net_yield_rate = (total_invoice_amount / total_charge_weight) if total_charge_weight else 0.0
                    
                    summary_data['Invoice AWB Count'] = invoice_awb_count
                    summary_data['Total Invoice Amount (Net Due)'] = total_invoice_amount
                    summary_data['Total Invoice Charge Weight'] = total_charge_weight
                    summary_data['Average Net Yield Rate'] = avg_net_yield_rate # Updated name/calculation
                    print(f"  -> Calculated Invoice Stats (using filtered df_awb_for_recon): Count={invoice_awb_count}, Amount={total_invoice_amount:.2f}, Weight={total_charge_weight:.2f}, Avg Rate={avg_net_yield_rate:.5f}")
                else:
                     print("  -> df_awb_for_recon became empty after re-applying 'Total' row filter for summary.")
                     summary_data['Invoice AWB Count'] = 0
                     summary_data['Total Invoice Amount (Net Due)'] = 0.0
                     summary_data['Total Invoice Charge Weight'] = 0.0
                     summary_data['Average Net Yield Rate'] = 0.0

            else:
                summary_data['Invoice AWB Count'] = 0

            # Calculate report totals from the reconciliation dataframe (which is based on left merge)
            if not df_reconciliation.empty and 'Net Due (Report)' in df_reconciliation.columns:
                # Filter df_reconciliation to remove the *newly added* total row before summing for summary
                df_rec_summary_input = df_reconciliation[df_reconciliation['AWB Prefix'] != 'Total'].copy()
                if not df_rec_summary_input.empty:
                    total_report_cost = df_rec_summary_input['Net Due (Report)'].sum()
                    difference_total_amount = total_report_cost - summary_data.get('Total Invoice Amount (Net Due)', 0.0)
                    summary_data['Total Report Amount (for Matched AWBs)'] = total_report_cost
                    summary_data['Difference (Report - Invoice)'] = difference_total_amount
                    print(f"  -> Calculated Report Stats (after filtering recon data): Total Cost={total_report_cost:.2f}, Difference={difference_total_amount:.2f}")
                else:
                    print("  -> Reconciliation data became empty after filtering 'Total' row for summary.")
                    summary_data['Total Report Amount (for Matched AWBs)'] = 0.0
                    summary_data['Difference (Report - Invoice)'] = 0.0 - summary_data.get('Total Invoice Amount (Net Due)', 0.0)
            else:
                summary_data['Total Report Amount (for Matched AWBs)'] = 0.0

            # Create DataFrame for summary
            df_summary = pd.DataFrame(list(summary_data.items()), columns=['Metric', 'Value'])
            
            # --- Write Sheets in Specified Order ---

            # 1. Summary Sheet
            sheet_name_summary = "Summary"
            df_summary.to_excel(writer, sheet_name=sheet_name_summary, index=False)
            print(f"  -> Written summary data to '{sheet_name_summary}' sheet.")
            worksheet_summary = writer.sheets[sheet_name_summary]
            # Autofit columns for summary sheet
            for column in worksheet_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_len = len(str(cell.value))
                        header_len = len(str(worksheet_summary[f"{column_letter}1"].value))
                        max_length = max(max_length, cell_len, header_len)
                    except: pass
                adjusted_width = (max_length + 2) * 1.1
                worksheet_summary.column_dimensions[column_letter].width = adjusted_width
            print(f"  -> Autofit columns for '{sheet_name_summary}'.")

            # 2. Reconciliation Sheet
            if not df_reconciliation.empty:
                sheet_name_rec = "Reconciliation"
                df_reconciliation.to_excel(writer, sheet_name=sheet_name_rec, index=False)
                print(f"  -> Written {len(df_reconciliation)} rows (incl. added total) to '{sheet_name_rec}' sheet.")
                worksheet_rec = writer.sheets[sheet_name_rec]
                # Formatting (Autofit, Table, Conditional)
                start_row_rec = 1
                end_row_rec = start_row_rec + len(df_reconciliation) # Includes newly added total row
                start_col_rec = 1
                end_col_rec = len(df_reconciliation.columns)
                # Adjust table range
                rec_table_range = f"{get_column_letter(start_col_rec)}{start_row_rec}:{get_column_letter(end_col_rec)}{end_row_rec}"
                print(f"  -> Calculated Reconciliation Table Range (incl. added total): {rec_table_range}")
                for column in worksheet_rec.columns:
                     max_length = 0
                     column_letter = column[0].column_letter
                     for cell in column:
                        try:
                            cell_len = len(str(cell.value))
                            header_len = len(str(worksheet_rec[f"{column_letter}1"].value))
                            max_length = max(max_length, cell_len, header_len)
                        except: pass
                     adjusted_width = (max_length + 2) * 1.1
                     worksheet_rec.column_dimensions[column_letter].width = adjusted_width
                print(f"  -> Autofit columns for '{sheet_name_rec}'.")
                tab_rec = Table(displayName="ReconciliationTable", ref=rec_table_range)
                style_rec = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_rec.tableStyleInfo = style_rec
                # Apply table if there are any rows (including the total row)
                if len(df_reconciliation) > 0:
                    worksheet_rec.add_table(tab_rec)
                    print(f"  -> Added Excel table formatting to '{sheet_name_rec}'.")
                # Apply conditional formatting if data rows exist (check > 1 because of added total row)
                discrepancy_col_letter = None
                col_letters = {} # Store column letters for specific fields
                for col_idx, value in enumerate(df_reconciliation.columns):
                    col_letter = get_column_letter(col_idx + 1)
                    col_letters[value] = col_letter # Store letter by column name
                    if value == 'Discrepancy Found':
                        discrepancy_col_letter = get_column_letter(col_idx + 1)
                        # Don't break, continue finding other column letters
                
                yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid') # Yellow for row highlight
                orange_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid') # Light red/orange for cell mismatches
                bright_red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid') # Bright Red for Diff Net Due
                black_thick_border = Border(left=Side(style='thick', color='000000'), # Changed color to black
                                          right=Side(style='thick', color='000000'), 
                                          top=Side(style='thick', color='000000'), 
                                          bottom=Side(style='thick', color='000000'))

                # Apply conditional formatting if data rows exist (check > 1 because of added total row)
                if discrepancy_col_letter and len(df_reconciliation) > 1:
                    # Apply conditional formatting only to data rows (excluding header and total row)
                    data_range_end_row = end_row_rec - 1 # Exclude the last (total) row
                    data_range_full_row = f"A2:{get_column_letter(end_col_rec)}{data_range_end_row}"
                    
                    # Check if data_range is valid (start row <= end row)
                    if data_range_end_row >= 2:
                        # 2. Highlight specific mismatching cells (apply before row highlight)
                        # Use ROUND for float comparison to handle precision issues.
                        # Check that both cells are not blank before comparing.
                        # Charge Weight Mismatch (e.g., cols C vs D)
                        if 'Charge Weight (Invoice)' in col_letters and 'Charge Weight (Report)' in col_letters:
                            cw_inv_col = col_letters['Charge Weight (Invoice)']
                            cw_rep_col = col_letters['Charge Weight (Report)']
                            formula_cw = [f'AND(ISNUMBER(${cw_inv_col}2), ISNUMBER(${cw_rep_col}2), ROUND(${cw_inv_col}2,2)<>ROUND(${cw_rep_col}2,2))']
                            # Apply to both columns C and D (adjust range)
                            try:
                                worksheet_rec.conditional_formatting.add(f'{cw_inv_col}2:{cw_inv_col}{data_range_end_row}', FormulaRule(formula=formula_cw, stopIfTrue=False, fill=orange_fill))
                                worksheet_rec.conditional_formatting.add(f'{cw_rep_col}2:{cw_rep_col}{data_range_end_row}', FormulaRule(formula=formula_cw, stopIfTrue=False, fill=orange_fill))
                                print(f"  -> Added CF for Charge Weight mismatch (Cols {cw_inv_col}, {cw_rep_col}).")
                            except Exception as cf_error:
                                print(f"  -> ERROR applying CF for Charge Weight: {cf_error}")
                            
                        # Net Yield Rate Mismatch (e.g., cols E vs F)
                        if 'Net Yield Rate (Invoice)' in col_letters and 'Net Yield Rate (Report)' in col_letters:
                            nyr_inv_col = col_letters['Net Yield Rate (Invoice)']
                            nyr_rep_col = col_letters['Net Yield Rate (Report)']
                            formula_nyr = [f'AND(ISNUMBER(${nyr_inv_col}2), ISNUMBER(${nyr_rep_col}2), ROUND(${nyr_inv_col}2,5)<>ROUND(${nyr_rep_col}2,5))'] # Higher precision for rate
                            # Apply to both columns E and F (adjust range)
                            try:
                                worksheet_rec.conditional_formatting.add(f'{nyr_inv_col}2:{nyr_inv_col}{data_range_end_row}', FormulaRule(formula=formula_nyr, stopIfTrue=False, fill=orange_fill))
                                worksheet_rec.conditional_formatting.add(f'{nyr_rep_col}2:{nyr_rep_col}{data_range_end_row}', FormulaRule(formula=formula_nyr, stopIfTrue=False, fill=orange_fill))
                                print(f"  -> Added CF for Net Yield Rate mismatch (Cols {nyr_inv_col}, {nyr_rep_col}).")
                            except Exception as cf_error:
                                print(f"  -> ERROR applying CF for Net Yield Rate: {cf_error}")

                        # Net Due Mismatch (e.g., cols G vs H, check Diff I != 0)
                        if 'Net Due (Invoice)' in col_letters and 'Net Due (Report)' in col_letters and 'Diff Net Due' in col_letters:
                            nd_inv_col = col_letters['Net Due (Invoice)']
                            nd_rep_col = col_letters['Net Due (Report)']
                            nd_diff_col = col_letters['Diff Net Due']
                            formula_nd_highlight_g_h = [f'AND(ISNUMBER(${nd_inv_col}2), ISNUMBER(${nd_rep_col}2), ROUND(${nd_diff_col}2,2)<>0)'] # For highlighting G & H
                            formula_nd_highlight_i = [f'ROUND(${nd_diff_col}2,2)<>0'] # For highlighting I (diff column itself)
                            # Apply to both columns G and H (adjust range)
                            try:
                                # Highlight G and H in light orange
                                worksheet_rec.conditional_formatting.add(f'{nd_inv_col}2:{nd_inv_col}{data_range_end_row}', FormulaRule(formula=formula_nd_highlight_g_h, stopIfTrue=False, fill=orange_fill))
                                worksheet_rec.conditional_formatting.add(f'{nd_rep_col}2:{nd_rep_col}{data_range_end_row}', FormulaRule(formula=formula_nd_highlight_g_h, stopIfTrue=False, fill=orange_fill))
                                print(f"  -> Added CF for Net Due mismatch (Cols {nd_inv_col}, {nd_rep_col}).")
                                # Highlight I (Diff Net Due) in bright red with border
                                worksheet_rec.conditional_formatting.add(f'{nd_diff_col}2:{nd_diff_col}{data_range_end_row}', FormulaRule(formula=formula_nd_highlight_i, stopIfTrue=False, fill=bright_red_fill, border=black_thick_border))
                                print(f"  -> Added CF for Diff Net Due itself (Col {nd_diff_col}).")
                            except Exception as cf_error:
                                 print(f"  -> ERROR applying CF for Net Due: {cf_error}")
                            
                        # 1. Apply Row Highlight Last (if Discrepancy Found is True)
                        formula = [f"${discrepancy_col_letter}2=TRUE"] # Formula still based on row 2
                        try:
                            worksheet_rec.conditional_formatting.add(data_range_full_row,
                                 FormulaRule(formula=formula, stopIfTrue=True, fill=yellow_fill))
                            print(f"  -> Added CF for Discrepancy Found row highlight (Range: {data_range_full_row}, Col: {discrepancy_col_letter}).")
                        except Exception as cf_error:
                             print(f"  -> ERROR applying CF for Discrepancy Found: {cf_error}")

            else:
                pd.DataFrame().to_excel(writer, sheet_name="Reconciliation", index=False)

            # 3. Invoices Sheet
            if not df_awb_final.empty:
                sheet_name_awb = "Invoices"
                df_awb_final.to_excel(writer, sheet_name=sheet_name_awb, index=False)
                print(f"  -> Written {len(df_awb_final)} rows to '{sheet_name_awb}' sheet.")
                worksheet_awb = writer.sheets[sheet_name_awb]
                # Formatting (Autofit, Table)
                start_row_awb = 1
                end_row_awb = start_row_awb + len(df_awb_final)
                start_col_awb = 1
                end_col_awb = len(df_awb_final.columns)
                ref_end_row_awb = end_row_awb 
                if 'Total' in df_awb_final['AWB Prefix'].values:
                   ref_end_row_awb = end_row_awb
                awb_table_range = f"{get_column_letter(start_col_awb)}{start_row_awb}:{get_column_letter(end_col_awb)}{ref_end_row_awb}"
                print(f"  -> Calculated AWB Table Range: {awb_table_range}")
                for column in worksheet_awb.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_len = len(str(cell.value))
                            header_len = len(str(worksheet_awb[f"{column_letter}1"].value))
                            max_length = max(max_length, cell_len, header_len)
                        except: pass
                    adjusted_width = (max_length + 2) * 1.1
                    worksheet_awb.column_dimensions[column_letter].width = adjusted_width
                print(f"  -> Autofit columns for '{sheet_name_awb}'.")
                tab_awb = Table(displayName="AWBTable", ref=awb_table_range)
                style_awb = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_awb.tableStyleInfo = style_awb
                if len(df_awb_final) > 0:
                    worksheet_awb.add_table(tab_awb)
                    print(f"  -> Added Excel table formatting to '{sheet_name_awb}'.")

                    # --- Apply Conditional Formatting for Net Due Discrepancies ---
                    if discrepancy_awbs:
                        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid') # Bright Red
                        # Find column indices for AWB Prefix and Serial
                        prefix_col_idx = -1
                        serial_col_idx = -1
                        header_row = worksheet_awb[1] # Header is row 1
                        for idx, cell in enumerate(header_row):
                            if cell.value == 'AWB Prefix':
                                prefix_col_idx = idx + 1 # 1-based index
                            elif cell.value == 'AWB Serial':
                                serial_col_idx = idx + 1 # 1-based index
                        
                        if prefix_col_idx != -1 and serial_col_idx != -1:
                            print(f"  -> Applying red highlight to invoice rows with Net Due discrepancies (Prefix Col: {prefix_col_idx}, Serial Col: {serial_col_idx})...")
                            # Iterate through data rows (starting from row 2, up to the row before total)
                            # end_row_awb already includes the total row if present
                            last_data_row = end_row_awb -1 if 'Total' in df_awb_final['AWB Prefix'].values else end_row_awb
                            for row_idx in range(2, last_data_row + 1):
                                try:
                                    prefix_val = worksheet_awb.cell(row=row_idx, column=prefix_col_idx).value
                                    serial_val = worksheet_awb.cell(row=row_idx, column=serial_col_idx).value
                                    # Ensure values are strings for comparison with the set
                                    awb_key = (str(prefix_val) if prefix_val is not None else '', 
                                               str(serial_val) if serial_val is not None else '')
                                    
                                    if awb_key in discrepancy_awbs:
                                        # Apply fill to all cells in this row within the table range
                                        for col_idx in range(start_col_awb, end_col_awb + 1):
                                            worksheet_awb.cell(row=row_idx, column=col_idx).fill = red_fill
                                except Exception as e:
                                    print(f"   -> Warning: Error applying highlight to row {row_idx}: {e}")
                            print(f"  -> Finished applying red highlight.")
                        else:
                            print("  -> Warning: Could not find AWB Prefix/Serial columns in Invoices sheet for highlighting.")
                    else:
                         print("  -> No Net Due discrepancies found, skipping invoice row highlighting.")
                    # --- End Conditional Formatting ---
                else:
                    print(f"  -> No data rows in '{sheet_name_awb}', skipping table creation.")
            else:
                pd.DataFrame().to_excel(writer, sheet_name="Invoices", index=False)
                print("  -> AWB DataFrame is empty. Written empty 'Invoices' sheet.")

            # 4. CCA Sheet
            if not df_cca_final.empty:
                sheet_name_cca = "CCA" 
                df_cca_final.to_excel(writer, sheet_name=sheet_name_cca, index=False)
                print(f"  -> Written {len(df_cca_final)} rows to '{sheet_name_cca}' sheet.")
                worksheet_cca = writer.sheets[sheet_name_cca]
                # Formatting (Autofit, Table)
                start_row_cca = 1
                end_row_cca = start_row_cca + len(df_cca_final)
                start_col_cca = 1
                end_col_cca = len(df_cca_final.columns)
                ref_end_row_cca = end_row_cca
                if 'Total' in df_cca_final['CCA Ref. No'].values:
                     ref_end_row_cca = end_row_cca
                cca_table_range = f"{get_column_letter(start_col_cca)}{start_row_cca}:{get_column_letter(end_col_cca)}{ref_end_row_cca}"
                print(f"  -> Calculated CCA Table Range: {cca_table_range}")
                for column in worksheet_cca.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_len = len(str(cell.value))
                            header_len = len(str(worksheet_cca[f"{column_letter}1"].value))
                            max_length = max(max_length, cell_len, header_len)
                        except: pass
                    adjusted_width = (max_length + 2) * 1.1
                    worksheet_cca.column_dimensions[column_letter].width = adjusted_width
                print(f"  -> Autofit columns for '{sheet_name_cca}'.")
                tab_cca = Table(displayName="CCATable", ref=cca_table_range)
                style_cca = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_cca.tableStyleInfo = style_cca
                if len(df_cca_final) > 0:
                    worksheet_cca.add_table(tab_cca)
                    print(f"  -> Added Excel table formatting to '{sheet_name_cca}'.")
                else:
                    print(f"  -> No data rows in '{sheet_name_cca}', skipping table creation.")
            else:
                pd.DataFrame().to_excel(writer, sheet_name="CCA", index=False)
                print("  -> CCA DataFrame is empty. Written empty 'CCA' sheet.")
            # --- End Writing Sheets ---
        
        print("--- Excel File Written Successfully ---")

    except Exception as e:
        print(f"Error writing Excel file: {e}")
        import traceback
        traceback.print_exc()
        # Return error state without a valid file path
        return {
             "invoices_rows": len(df_awb_final) -1 if not df_awb_final.empty and 'Total' in df_awb_final['AWB Prefix'].values else (len(df_awb_final) if not df_awb_final.empty else 0),
             "cca_rows": len(df_cca_final) -1 if not df_cca_final.empty and 'Total' in df_cca_final['CCA Ref. No'].values else (len(df_cca_final) if not df_cca_final.empty else 0),
             "excel_file": "", 
             "download_url": "",
             "total_net_due_awb": total_net_due_awb # Return calculated total even if write fails
        }

    # Adjust row counts to exclude total row before returning
    invoices_rows_count = 0
    if not df_awb_final.empty:
         invoices_rows_count = len(df_awb_final)
         if 'Total' in df_awb_final['AWB Prefix'].values:
             invoices_rows_count -= 1
             
    cca_rows_count = 0
    if not df_cca_final.empty:
        cca_rows_count = len(df_cca_final)
        if 'Total' in df_cca_final['CCA Ref. No'].values:
            cca_rows_count -= 1

    # Return results including the CCA count
    return {
        "invoices_rows": invoices_rows_count,
        "cca_rows": cca_rows_count,
        "excel_file": excel_filename,
        "download_url": download_url,
        "total_net_due_awb": total_net_due_awb 
    }


@app.route("/upload", methods=["POST"])
def upload():
    if 'file' not in request.files or 'report_file' not in request.files:
        flash('Both Invoice/CCA file and Report file are required.', 'error')
        return redirect(url_for('index'))

    invoice_file = request.files['file']
    report_file = request.files['report_file']

    if invoice_file.filename == '' or report_file.filename == '':
        flash('No selected file for one or both inputs.', 'error')
        return redirect(url_for('index'))

    # --- Process Invoice File ---
    invoice_file_path = None # Initialize
    if invoice_file and allowed_file(invoice_file.filename, ALLOWED_EXTENSIONS):
        invoice_filename = secure_filename(invoice_file.filename)
        invoice_file_path = os.path.join(app.config["UPLOAD_FOLDER"], invoice_filename)
        try:
            invoice_file.save(invoice_file_path)
            print(f"Invoice file saved to: {invoice_file_path}")
        except Exception as e:
            flash(f"Error saving invoice file: {e}", 'error')
            return redirect(url_for('index'))
    else:
        flash(f'Invalid invoice file type. Allowed: {", ".join(ALLOWED_EXTENSIONS)}', 'error')
        return redirect(url_for('index'))

    # --- Process Report File ---
    df_report = None
    report_file_path = None
    if report_file and allowed_file(report_file.filename, ALLOWED_REPORT_EXTENSIONS):
        report_filename = secure_filename(report_file.filename)
        report_file_path = os.path.join(app.config["UPLOAD_FOLDER"], report_filename)
        try:
            report_file.save(report_file_path)
            print(f"Report file saved to: {report_file_path}")
            # Read the saved report .xls file, specifying header row and data types
            print(f"Reading report file: {report_file_path} with header on row 8 (index 7)")
            df_report = pd.read_excel(
                report_file_path,
                engine='xlrd',
                header=7, # Assumes header is on the 8th row (0-indexed 7)
                dtype={'awbprefix': str, 'awbsuffix': str} # Specify dtype for merge keys
            )
            print(f"Report file read successfully. Shape: {df_report.shape}")

            # Clean column names by stripping whitespace and lowercasing
            df_report.columns = df_report.columns.str.strip().str.lower()
            print(f"Cleaned report columns (lower, strip): {df_report.columns.tolist()}")

            # --- Updated Validation for required columns ---
            required_report_cols = ['awbprefix', 'awbsuffix', 'chargewt', 'frt_cost_rate', 'total_cost']
            missing_cols = [col for col in required_report_cols if col not in df_report.columns]

            if missing_cols:
                 flash(f"Report file is missing one or more required columns (checked names, lowercased & stripped: {', '.join(missing_cols)}). Ensure header is on row 8 and columns exist. Actual found: {df_report.columns.tolist()}", 'error')
                 # Clean up invoice file before redirecting
                 if invoice_file_path and os.path.exists(invoice_file_path): os.remove(invoice_file_path)
                 if report_file_path and os.path.exists(report_file_path): os.remove(report_file_path) # Clean up report file too
                 return redirect(url_for('index'))

        except Exception as e:
            flash(f"Error reading or processing report file: {e}", 'error')
            if invoice_file_path and os.path.exists(invoice_file_path): os.remove(invoice_file_path)
            # Clean up report file if it exists
            if report_file_path and os.path.exists(report_file_path):
                try: os.remove(report_file_path)
                except Exception as rm_e: print(f"Error removing report file during error handling: {rm_e}")
            return redirect(url_for('index'))
    else:
        flash(f'Invalid report file type. Allowed: {", ".join(ALLOWED_REPORT_EXTENSIONS)}', 'error')
        if invoice_file_path and os.path.exists(invoice_file_path): os.remove(invoice_file_path) # Clean up invoice file
        return redirect(url_for('index'))

    # --- Process the invoice file, passing the report data ---
    try:
        results = process_file(invoice_file_path, report_data_df=df_report)

        # Check for processing errors indicated by missing excel_file key
        if not results.get("excel_file"):
             # Assuming errors are flashed within process_file or handled by empty df checks
             flash(f"Processing completed, but encountered issues. Excel file not generated.", 'warning') # Generic message if needed
             # Error/Warning should have been flashed already, redirect
             return redirect(url_for('index'))


        if results["invoices_rows"] == 0 and results["cca_rows"] == 0: # Check both
            flash(f"Processing Complete: No AWB or CCA data found in '{invoice_filename}'. Excel file not generated.", 'warning')
        else:
            # Display success and results using flash messages
            flash(f"File '{invoice_filename}' processed successfully!")
            flash(f"AWB Rows Processed: {results['invoices_rows']}")
            flash(f"CCA Rows Processed: {results['cca_rows']}") # Updated CCA count name
            flash(f"Total Net Due for AWB (from Invoice): {results['total_net_due_awb']:.2f}")
            flash(f"<a href='{results['download_url']}'>Download Processed Excel File: {results['excel_file']}</a>", 'download') # Use a specific category for styling

        return redirect(url_for('index'))

    except Exception as e:
        error_message = f"Error processing file: {e}"
        print(error_message)
        import traceback
        traceback.print_exc()
        flash(f"Error during processing: {error_message}", 'error')
        return redirect(url_for('index'))

    finally:
        # Clean up uploaded files
        if invoice_file_path and os.path.exists(invoice_file_path):
            try: os.remove(invoice_file_path)
            except Exception as e: print(f"Error removing invoice file {invoice_file_path}: {e}")
        if report_file_path and os.path.exists(report_file_path):
             try: os.remove(report_file_path)
             except Exception as e: print(f"Error removing report file {report_file_path}: {e}")

if __name__ == '__main__':
    # Set host to '0.0.0.0' to make it accessible on the network
    app.run(debug=True, host='0.0.0.0') 